import sqlite3
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
import os
import time
import re
import threading
from datetime import datetime
from functools import wraps
from werkzeug.utils import secure_filename
from urllib.parse import quote
import secrets
from dotenv import load_dotenv

# Load environment variables from .env
load_dotenv()

app = Flask(__name__)
app.secret_key = 'pikr_serasi_secret_key'

# --- API CONFIGURATION ---
# API Key untuk proteksi endpoint REST API
# Ganti nilai ini atau gunakan environment variable: export PIKR_API_KEY="key-anda"
API_KEY = os.environ.get('PIKR_API_KEY', 'pikr-admin-secret-2024')

# --- DATA ACCESS LAYER (DAL) ---
DB_PATH = 'database/pikr.db'
UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Simple mapping for counselor -> whatsapp number (no plus, country code first)
COUNSELOR_WHATSAPP = {
    'konselor_sebaya': '6289525165373'
}

# --- DATABASE WRAPPER FOR POSTGRES (SUPABASE) AND SQLITE ---
class PIKRDatabaseConnection:
    def __init__(self, db_path=None, supabase_url=None):
        self.supabase_url = supabase_url
        self.db_path = db_path or 'database/pikr.db'
        self.modified_tables = set()

        if self.supabase_url:
            import psycopg2
            from psycopg2.extras import DictCursor
            # Connect to PostgreSQL/Supabase
            self.conn = psycopg2.connect(self.supabase_url)
            self.cursor = self.conn.cursor(cursor_factory=DictCursor)
        else:
            # Connect to SQLite
            self.conn = sqlite3.connect(self.db_path)
            self.conn.row_factory = sqlite3.Row
            self.cursor = self.conn.cursor()

    def execute(self, query, params=None):
        # Detect if this query writes to database
        query_upper = query.strip().upper()
        is_write = any(query_upper.startswith(op) for op in ["INSERT", "UPDATE", "DELETE"])
        
        if is_write:
            table_name = None
            words = query_upper.split()
            if "INSERT INTO" in query_upper and len(words) > 2:
                idx = words.index("INTO")
                if idx != -1 and idx + 1 < len(words):
                    table_name = words[idx+1].split('(')[0].strip().strip('"').strip("'").lower()
            elif "UPDATE" in query_upper and len(words) > 1:
                table_name = words[1].strip().strip('"').strip("'").lower()
            elif "DELETE FROM" in query_upper and len(words) > 2:
                idx = words.index("FROM")
                if idx != -1 and idx + 1 < len(words):
                    table_name = words[idx+1].strip().strip('"').strip("'").lower()
            
            if table_name:
                self.modified_tables.add(table_name)

        if self.supabase_url:
            # Convert SQLite placeholders '?' to '%s'
            query = query.replace('?', '%s')
            # Convert SQLite double quoted string literals to single quotes
            query = re.sub(r'"([^"]*)"', r"'\1'", query)
            
            if params is not None:
                if not isinstance(params, (tuple, list)):
                    params = (params,)
                self.cursor.execute(query, params)
            else:
                self.cursor.execute(query)
        else:
            if params is not None:
                self.cursor.execute(query, params)
            else:
                self.cursor.execute(query)
        return self

    def fetchone(self):
        row = self.cursor.fetchone()
        if row is None:
            return None
        return row

    def fetchall(self):
        return self.cursor.fetchall()

    def commit(self):
        self.conn.commit()

    def close(self):
        self.cursor.close()
        self.conn.close()
        
        # Trigger async sheets sync if any tables were modified
        if self.modified_tables:
            for table in self.modified_tables:
                trigger_async_sheets_sync(table)

    def __iter__(self):
        return iter(self.cursor)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is not None:
            self.conn.rollback()
        else:
            self.commit()
        self.close()

def sync_worker(table_name):
    try:
        supabase_url = os.environ.get('SUPABASE_DB_URL')
        if supabase_url:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(supabase_url)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute(f"SELECT * FROM {table_name}")
            rows = [dict(r) for r in cursor.fetchall()]
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute(f"SELECT * FROM {table_name}")
            rows = [dict(r) for r in cursor.fetchall()]
            cursor.close()
            conn.close()
            
        from sheets_sync import sync_table_to_sheets
        sync_table_to_sheets(table_name, rows)
    except Exception as e:
        print(f"Error in sheet sync worker for {table_name}: {e}")

def trigger_async_sheets_sync(table_name):
    t = threading.Thread(target=sync_worker, args=(table_name,))
    t.daemon = True
    t.start()

def get_db_connection():
    supabase_url = os.environ.get('SUPABASE_DB_URL')
    return PIKRDatabaseConnection(db_path=DB_PATH, supabase_url=supabase_url)

# --- BUSINESS LOGIC LAYER (Service Layer) ---
def add_notification(username, message, link):
    conn = get_db_connection()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    conn.execute('INSERT INTO notifications (username, message, link, timestamp) VALUES (?, ?, ?, ?)',
                 (username, message, link, timestamp))
    conn.commit()
    conn.close()

def validate_age(birth_date_str):
    try:
        birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d')
        today = datetime.today()
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        return 10 <= age <= 24
    except ValueError:
        return False

def notify_role(role, message, link):
    """Mengirim notifikasi ke semua user dengan role tertentu"""
    conn = get_db_connection()
    users = conn.execute('SELECT username FROM users WHERE role = ?', (role,)).fetchall()
    conn.close()
    
    current_user = session.get('username')
    for u in users:
        if u['username'] != current_user:
            add_notification(u['username'], message, link)

def notify_admins(message, link):
    """Mengirim notifikasi khusus ke semua Admin"""
    notify_role('admin_pikr', message, link)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash("Silakan login terlebih dahulu", "warning")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def api_key_required(f):
    """Decorator untuk proteksi endpoint REST API menggunakan X-API-Key header"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        key = request.headers.get('X-API-Key') or request.args.get('api_key')
        if not key or key != API_KEY:
            return jsonify({"success": False, "message": "Unauthorized: API key tidak valid atau tidak ada."}), 401
        return f(*args, **kwargs)
    return decorated_function

def api_success(data=None, message="Berhasil", status=200):
    """Helper untuk standarisasi response sukses API"""
    resp = {"success": True, "message": message}
    if data is not None:
        resp["data"] = data
    return jsonify(resp), status

def api_error(message="Terjadi kesalahan", status=400):
    """Helper untuk standarisasi response error API"""
    return jsonify({"success": False, "message": message}), status

# --- PRESENTATION LAYER (Controller/Routes) ---

@app.route('/api/notifications')
@login_required
def get_notifications():
    conn = get_db_connection()
    notifs = conn.execute('SELECT * FROM notifications WHERE username = ? ORDER BY id DESC', (session.get('username'),)).fetchall()
    conn.close()
    return jsonify([dict(n) for n in notifs])

@app.route('/api/notifications/read/<int:notif_id>', methods=['POST'])
@login_required
def read_notification(notif_id):
    conn = get_db_connection()
    conn.execute('UPDATE notifications SET is_read = 1 WHERE id = ? AND username = ?', (notif_id, session.get('username')))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/notifications/read_all', methods=['POST'])
@login_required
def read_all_notifications():
    conn = get_db_connection()
    conn.execute('UPDATE notifications SET is_read = 1 WHERE username = ?', (session.get('username'),))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/notifications/clear', methods=['POST'])
@login_required
def clear_notifications():
    conn = get_db_connection()
    conn.execute('DELETE FROM notifications WHERE username = ?', (session.get('username'),))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/')
def index():
    conn = get_db_connection()
    latest_ach = conn.execute('SELECT * FROM achievements ORDER BY id DESC LIMIT 1').fetchone()
    conn.close()
    
    featured_ach = None
    if latest_ach:
        featured_ach = dict(latest_ach)
        if featured_ach.get('dokumen'):
            featured_ach['foto_list'] = [f.strip() for f in featured_ach['dokumen'].split(',') if f.strip()]
        else:
            featured_ach['foto_list'] = []
            
    return render_template('index.html', featured_ach=featured_ach)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        birth_date = request.form['birth_date']
        role = request.form['role']

        if not validate_age(birth_date):
            flash("Pendaftaran Gagal: Usia harus antara 10-24 tahun.", "danger")
            return redirect(url_for('register'))

        conn = get_db_connection()
        existing_user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        
        if existing_user:
            conn.close()
            flash("Username sudah digunakan!", "danger")
            return redirect(url_for('register'))

        conn.execute('INSERT INTO users (username, password, birth_date, role) VALUES (?, ?, ?, ?)',
                     (username, password, birth_date, role))
        conn.commit()
        conn.close()
        
        flash(f"Registrasi sebagai {role} berhasil!", "success")
        return redirect(url_for('login'))

    return render_template('register.html')

def authenticate_user(username, password):
    conn = get_db_connection()
    user = conn.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password)).fetchone()
    conn.close()
    return dict(user) if user else None

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = authenticate_user(request.form['username'], request.form['password'])
        if user:
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['username'] = user['username']
            return redirect(url_for('dashboard'))
        flash("Login gagal!", "danger")
    return render_template('login.html')

def update_user_data(username, new_data):
    if not new_data:
        return False
    conn = get_db_connection()
    set_clause = ', '.join([f"{k} = ?" for k in new_data.keys()])
    values = list(new_data.values())
    values.append(username)
    conn.execute(f'UPDATE users SET {set_clause} WHERE username = ?', values)
    conn.commit()
    conn.close()
    return True

def add_points(username, points):
    conn = get_db_connection()
    conn.execute('UPDATE users SET points = points + ? WHERE username = ?', (points, username))
    conn.commit()
    conn.close()
    return True

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    my_username = session.get('username')
    conn = get_db_connection()
    user_row = conn.execute('SELECT * FROM users WHERE username = ?', (my_username,)).fetchone()
    conn.close()
    user = dict(user_row) if user_row else None
            
    if request.method == 'POST':
        new_bio = request.form.get('bio')
        new_password = request.form.get('password')
        
        updates = {"bio": new_bio}
        if new_password:
            updates["password"] = new_password
            
        # Handle Profile Picture
        file = request.files.get('profile_pic')
        if file and file.filename:
            filename = f"profile_{my_username}_{int(time.time())}.png"
            filepath = os.path.join('static/uploads/profiles', filename)
            os.makedirs('static/uploads/profiles', exist_ok=True)
            file.save(filepath)
            updates["profile_pic"] = filepath
            
        update_user_data(my_username, updates)
        flash('Profil berhasil diperbarui!', 'success')
        return redirect(url_for('profile'))
        
    return render_template('profile.html', user=user)

@app.route('/dashboard')
@login_required
def dashboard():
    my_username = session['username']
    my_role = session.get('role')
    conn = get_db_connection()
    
    total_users = conn.execute('SELECT COUNT(*) as c FROM users').fetchone()['c']
    total_konselor = conn.execute('SELECT COUNT(*) as c FROM users WHERE role = "konselor"').fetchone()['c']
    total_remaja = conn.execute('SELECT COUNT(*) as c FROM users WHERE role = "anggota_remaja"').fetchone()['c']
    
    recent_users = conn.execute('SELECT * FROM users ORDER BY id DESC LIMIT 5').fetchall()
    total_articles = conn.execute('SELECT COUNT(*) as c FROM education').fetchone()['c']
    
    chat_partners = [row['partner'] for row in conn.execute('''
        SELECT receiver as partner FROM messages WHERE sender = ?
        UNION
        SELECT sender as partner FROM messages WHERE receiver = ?
    ''', (my_username, my_username)).fetchall()]

    if my_role == 'admin_pikr':
        my_sessions = conn.execute('SELECT * FROM sessions').fetchall()
    elif my_role == 'konselor':
        my_sessions = conn.execute('SELECT * FROM sessions WHERE counselor_name = ?', (my_username,)).fetchall()
    elif my_role == 'klinik_kesehatan':
        my_sessions = conn.execute('SELECT * FROM sessions WHERE priority IN ("emergency", "high")').fetchall()
    else:
        my_sessions = conn.execute('SELECT * FROM sessions WHERE member_name = ?', (my_username,)).fetchall()

    pending = sum(1 for s in my_sessions if s['status'].lower() == 'pending')
    approved = sum(1 for s in my_sessions if s['status'].lower() == 'approved')
    rejected = sum(1 for s in my_sessions if s['status'].lower() == 'rejected')
    chart_data = [pending, approved, rejected]
    
    emergency_sessions = conn.execute('SELECT * FROM sessions WHERE priority IN ("emergency", "high")').fetchall()
    total_emergency = len(emergency_sessions)

    joined_events = conn.execute('''
        SELECT e.* FROM events e 
        JOIN event_participants p ON e.id = p.event_id 
        WHERE p.username = ?
    ''', (my_username,)).fetchall()

    user_row = conn.execute('SELECT points FROM users WHERE username = ?', (my_username,)).fetchone()
    user_points = user_row['points'] if user_row else 0
    
    recent_posts = conn.execute('SELECT * FROM forum_posts ORDER BY id DESC LIMIT 5').fetchall()
    konselor_data = conn.execute('SELECT * FROM users WHERE role = "konselor"').fetchall()
    
    conn.close()

    return render_template('dashboard.html', 
                           role=my_role, 
                           chart_data=chart_data,
                           total_users=total_users,
                           total_remaja=total_remaja,
                           total_running_sessions=chart_data[1],
                           total_konselor=total_konselor,
                           total_articles=total_articles,
                           recent_users=[dict(u) for u in recent_users],
                           recent_posts=[dict(p) for p in recent_posts],
                           konselor_data=[dict(k) for k in konselor_data],
                           chat_partners=chat_partners,
                           emergency_sessions=[dict(s) for s in emergency_sessions],
                           total_emergency=total_emergency,
                           joined_events=[dict(e) for e in joined_events],
                           user_points=user_points)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

# --- FORUM FITUR (DENGAN COCH BOX PENGUMUMAN) ---
@app.route('/forum', methods=['GET', 'POST'])
def forum():
    conn = get_db_connection()
    if request.method == 'POST':
        if 'user_id' not in session:
            flash("Silakan login untuk membuat posting forum.", "warning")
            conn.close()
            return redirect(url_for('login'))

        content = request.form.get('content')
        is_announcement = 1 if request.form.get('is_announcement') == 'on' else 0
        
        if content and content.strip():
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            conn.execute('INSERT INTO forum_posts (username, content, role, is_announcement, date) VALUES (?, ?, ?, ?, ?)',
                         (session.get('username'), content, session.get('role'), is_announcement, date_str))
            conn.commit()

            if is_announcement:
                users = conn.execute('SELECT username FROM users').fetchall()
                for u in users:
                    if u['username'] != session.get('username'):
                        add_notification(u['username'], f"📢 Pengumuman: {content[:30]}...", url_for('forum'))
            else:
                notify_role('konselor', f"📝 Kiriman forum baru dari {session['username']}", url_for('forum'))
                notify_admins(f"📝 Kiriman forum baru dari {session['username']}", url_for('forum'))
            
            valid_users = [u['username'] for u in conn.execute('SELECT username FROM users').fetchall()]
            for username in valid_users:
                if f"@{username.lower()}" in content.lower() and username != session.get('username'):
                    add_notification(username, f"💬 Kamu di-tag oleh {session['username']} di forum", url_for('forum'))
                        
            flash('Berhasil mengirim ke forum!', 'success')
        conn.close()
        return redirect(url_for('forum'))
        
    posts = conn.execute('SELECT * FROM forum_posts ORDER BY is_announcement DESC, id DESC').fetchall()
    valid_usernames = [u['username'] for u in conn.execute('SELECT username FROM users').fetchall()]
    conn.close()
    return render_template('forum.html', posts=[dict(p) for p in posts], valid_usernames=valid_usernames)

@app.route('/delete_post/<int:post_id>')
@login_required
def delete_post(post_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM forum_posts WHERE id = ?', (post_id,))
    conn.commit()
    conn.close()
    flash('Postingan berhasil dihapus!', 'success')
    return redirect(url_for('forum'))

@app.route('/edit_post/<int:post_id>', methods=['POST'])
@login_required
def edit_post(post_id):
    new_content = request.form.get('content')
    conn = get_db_connection()
    conn.execute('UPDATE forum_posts SET content = ? WHERE id = ?', (new_content, post_id))
    conn.commit()
    conn.close()
    flash('Postingan berhasil diperbarui!', 'success')
    return redirect(url_for('forum'))

# --- KONSELING FITUR ---
@app.route('/konseling', methods=['GET', 'POST'])
def counseling():
    conn = get_db_connection()
    if request.method == 'POST':
        counselor_name = request.form.get('counselor_name')
        topic = request.form.get('topic')
        date = request.form.get('date')
        time = request.form.get('time')
        mode = request.form.get('mode')

        is_guest = 'user_id' not in session
        if is_guest:
            consent = request.form.get('consent')
            nickname = request.form.get('nickname')
            if not consent:
                flash('Anda harus menyetujui persetujuan untuk melanjutkan.', 'warning')
                conn.close()
                return redirect(url_for('counseling'))
            member_name = nickname.strip() if nickname and nickname.strip() else f"Guest-{secrets.token_hex(3)}"
        else:
            member_name = session.get('username')

        if not counselor_name:
            flash('Silakan pilih konselor terlebih dahulu.', 'warning')
            conn.close()
            return redirect(url_for('counseling'))

        is_conflict = conn.execute('''
            SELECT 1 FROM sessions 
            WHERE counselor_name = ? AND date = ? AND time = ? AND status IN ('PENDING', 'APPROVED')
        ''', (counselor_name, date, time)).fetchone()

        if is_conflict:
            flash(f"Maaf, jadwal pada tanggal {date} jam {time} dengan konselor {counselor_name} sudah dibooking. Silakan pilih waktu lain.", "danger")
            conn.close()
            return redirect(url_for('counseling'))

        conn.execute('''
            INSERT INTO sessions (member_name, counselor_name, topic, date, time, status) 
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (member_name, counselor_name, topic, date, time, "PENDING"))
        conn.commit()

        if is_guest:
            add_notification(counselor_name, f"📅 Permintaan konseling baru (Guest) dari {member_name}", url_for('counseling'))
            notify_admins(f"📅 Permintaan konseling baru (Guest) dari {member_name} untuk {counselor_name}", url_for('counseling'))
        else:
            add_notification(counselor_name, f"📅 Permintaan konseling baru dari {member_name}", url_for('counseling'))
            notify_admins(f"📅 Permintaan konseling baru dari {member_name} untuk {counselor_name}", url_for('counseling'))

        counselor_key = (counselor_name or '').strip().lower()
        wa_number = COUNSELOR_WHATSAPP.get(counselor_key)
        conn.close()

        if mode and mode.lower() == 'online' and wa_number:
            message = f"Halo {counselor_name}, saya {'anonim' if is_guest else member_name} ingin konfirmasi sesi konseling pada {date} jam {time} (online). Mohon bantuannya."
            wa_url = f"https://wa.me/{wa_number}?text={quote(message)}"
            return redirect(wa_url)

        flash("Permintaan konseling berhasil dikirim!", "success")
        return redirect(url_for('counseling'))

    if session.get('user_id'):
        if session.get('role') == 'admin_pikr':
            user_sessions = conn.execute('SELECT * FROM sessions ORDER BY id DESC').fetchall()
        elif session.get('role') == 'klinik_kesehatan':
            user_sessions = conn.execute('SELECT * FROM sessions WHERE priority IN ("emergency", "high") ORDER BY id DESC').fetchall()
        else:
            user_sessions = conn.execute('SELECT * FROM sessions WHERE member_name = ? OR counselor_name = ? ORDER BY id DESC', 
                                         (session.get('username'), session.get('username'))).fetchall()
    else:
        user_sessions = []

    konselor_data = conn.execute('SELECT * FROM users WHERE role = "konselor"').fetchall()
    conn.close()
    return render_template('counseling.html', 
                           sessions=[{**dict(s), 'session_id': s['id']} for s in user_sessions], 
                           konselor_data=[dict(k) for k in konselor_data],
                           is_guest=('user_id' not in session))


@app.route('/konseling_guest', methods=['GET', 'POST'])
def counseling_guest():
    return redirect(url_for('counseling'))

@app.route('/approve_session/<int:session_id>')
@login_required
def approve_session(session_id):
    if session.get('role') not in ['konselor', 'admin_pikr']:
        flash("Akses ditolak", "danger")
        return redirect(url_for('counseling'))
        
    conn = get_db_connection()
    s = conn.execute('SELECT * FROM sessions WHERE id = ?', (session_id,)).fetchone()
    if s:
        conn.execute('UPDATE sessions SET status = "APPROVED" WHERE id = ?', (session_id,))
        conn.commit()
        add_notification(s['member_name'], "✅ Jadwal konseling Anda disetujui!", url_for('counseling'))
        notify_admins(f"✅ Konselor {session['username']} menyetujui sesi #{session_id}", url_for('counseling'))
        flash("Sesi disetujui!", "success")
    conn.close()
    return redirect(url_for('counseling'))

@app.route('/reject_session/<int:session_id>')
@login_required
def reject_session(session_id):
    if session.get('role') not in ['konselor', 'admin_pikr']:
        flash("Akses ditolak", "danger")
        return redirect(url_for('counseling'))
        
    conn = get_db_connection()
    s = conn.execute('SELECT * FROM sessions WHERE id = ?', (session_id,)).fetchone()
    if s:
        conn.execute('UPDATE sessions SET status = "REJECTED" WHERE id = ?', (session_id,))
        conn.commit()
        add_notification(s['member_name'], "❌ Jadwal konseling Anda ditolak.", url_for('counseling'))
        notify_admins(f"❌ Konselor {session['username']} menolak sesi #{session_id}", url_for('counseling'))
        flash("Sesi ditolak!", "warning")
    conn.close()
    return redirect(url_for('counseling'))

@app.route('/delete_session/<int:session_id>')
@login_required
def delete_session(session_id):
    conn = get_db_connection()
    s = conn.execute('SELECT * FROM sessions WHERE id = ?', (session_id,)).fetchone()
    if s and (session.get('role') == 'admin_pikr' or s['member_name'] == session['username'] or s['counselor_name'] == session['username']):
        conn.execute('DELETE FROM sessions WHERE id = ?', (session_id,))
        conn.commit()
        flash("Riwayat sesi dihapus", "success")
    else:
        flash("Gagal menghapus atau akses ditolak", "danger")
    conn.close()
    return redirect(url_for('counseling'))

@app.route('/edit_session/<int:session_id>', methods=['POST'])
@login_required
def edit_session(session_id):
    conn = get_db_connection()
    s = conn.execute('SELECT * FROM sessions WHERE id = ?', (session_id,)).fetchone()
    if s and (session.get('role') == 'admin_pikr' or s['member_name'] == session['username'] or s['counselor_name'] == session['username']):
        new_date = request.form.get('date')
        new_time = request.form.get('time')
        new_status = request.form.get('status', s['status']) if session.get('role') in ['admin_pikr', 'konselor'] else s['status']
        
        conn.execute('UPDATE sessions SET date = ?, time = ?, status = ? WHERE id = ?', 
                     (new_date, new_time, new_status, session_id))
        conn.commit()
        flash("Jadwal diperbarui!", "success")
    else:
        flash("Gagal memperbarui jadwal atau akses ditolak", "danger")
    conn.close()
    return redirect(url_for('counseling'))

# --- EDUKASI & EVENT TERPADU ---
@app.route('/edukasi-event')
def education_event():
    conn = get_db_connection()
    # 1. Fetch Education Articles & comments
    articles_raw = conn.execute('SELECT * FROM education ORDER BY id DESC').fetchall()
    articles = []
    for art in articles_raw:
        a = dict(art)
        a['comments'] = [dict(c) for c in conn.execute('SELECT * FROM education_comments WHERE article_id = ?', (a['id'],)).fetchall()]
        articles.append(a)
        
    # 2. Fetch Events & participants
    events_raw = conn.execute('SELECT * FROM events ORDER BY date DESC').fetchall()
    all_events = []
    for e in events_raw:
        ev = dict(e)
        participants = conn.execute('SELECT username FROM event_participants WHERE event_id = ?', (e['id'],)).fetchall()
        ev['participants'] = [p['username'] for p in participants]
        if ev.get('foto'):
            ev['foto_list'] = [f.strip() for f in ev['foto'].split(',') if f.strip()]
        else:
            ev['foto_list'] = []
        all_events.append(ev)
        
    conn.close()
    
    active_tab = request.args.get('tab', 'edukasi')
    return render_template('education_event.html', articles=articles, events=all_events, active_tab=active_tab)

@app.route('/edukasi')
def education():
    conn = get_db_connection()
    articles_raw = conn.execute('SELECT * FROM education ORDER BY id DESC').fetchall()
    articles = []
    for art in articles_raw:
        a = dict(art)
        a['comments'] = [dict(c) for c in conn.execute('SELECT * FROM education_comments WHERE article_id = ?', (a['id'],)).fetchall()]
        articles.append(a)
    conn.close()
    return render_template('education.html', articles=articles)

@app.route('/event')
def event():
    return redirect(url_for('events'))

@app.route('/event/<int:event_id>')
def event_detail(event_id):
    conn = get_db_connection()
    ev_row = conn.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    if not ev_row:
        conn.close()
        flash('Kegiatan tidak ditemukan.', 'danger')
        return redirect(url_for('education_event', tab='event'))
    ev = dict(ev_row)
    if ev.get('foto'):
        ev['foto_list'] = [f.strip() for f in ev['foto'].split(',') if f.strip()]
    else:
        ev['foto_list'] = []
    participants = conn.execute('SELECT username FROM event_participants WHERE event_id = ?', (event_id,)).fetchall()
    ev['participants'] = [p['username'] for p in participants]
    conn.close()
    return render_template('event_detail.html', ev=ev)

# --- INOVASI & PROGRAM KERJA ---

# List innovations
@app.route('/inovasi-proker')
def innovations():
    conn = get_db_connection()
    innovations_raw = conn.execute('SELECT * FROM innovations ORDER BY id DESC').fetchall()
    innovations = []
    for inn in innovations_raw:
        i = dict(inn)
        i['comments'] = [dict(c) for c in conn.execute('SELECT * FROM innovations_comments WHERE innovation_id = ?', (i['id'],)).fetchall()]
        innovations.append(i)
    conn.close()
    return render_template('innovations.html', innovations=innovations)

# Innovation detail
@app.route('/inovasi-proker/<int:inn_id>')
def innovation_detail(inn_id):
    conn = get_db_connection()
    inn = conn.execute('SELECT * FROM innovations WHERE id = ?', (inn_id,)).fetchone()
    if inn:
        innovation = dict(inn)
        innovation['comments'] = [dict(c) for c in conn.execute('SELECT * FROM innovations_comments WHERE innovation_id = ?', (inn_id,)).fetchall()]
        conn.close()
        return render_template('innovation_detail.html', innovation=innovation)
    conn.close()
    flash('Inovasi tidak ditemukan', 'warning')
    return redirect(url_for('innovations'))

# Add innovation (admin/konselor)
@app.route('/add_innovation', methods=['POST'])
@login_required
def add_innovation():
    if session.get('role') not in ['konselor', 'admin_pikr']:
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('innovations'))
    title = request.form.get('title')
    content = request.form.get('content', '')
    file_dokumen = request.files.get('dokumen')
    nama_dokumen = None
    if file_dokumen and file_dokumen.filename != '':
        nama_dokumen = secure_filename(file_dokumen.filename)
        nama_dokumen = f"{int(datetime.now().timestamp())}_{nama_dokumen}"
        file_dokumen.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen))
    if (not content or content.strip() == '') and not nama_dokumen:
        flash('Gagal menerbitkan: Mohon isi teks atau lampirkan file dokumen.', 'danger')
        return redirect(url_for('innovations'))
    conn = get_db_connection()
    conn.execute('INSERT INTO innovations (title, content, author, author_id, dokumen) VALUES (?, ?, ?, ?, ?)',
                 (title, content, session['username'], session['user_id'], nama_dokumen))
    conn.commit()
    conn.close()
    notify_role('anggota_remaja', f"💡 Inovasi Baru: {title}", url_for('innovations'))
    notify_admins(f"💡 Inovasi baru diterbitkan oleh {session['username']}: {title}", url_for('innovations'))
    flash('Inovasi berhasil diterbitkan!', 'success')
    return redirect(url_for('innovations'))

# Admin wrapper
@app.route('/admin/innovation/add', methods=['POST'])
@login_required
def add_innovation_admin():
    return add_innovation()

# Delete innovation
@app.route('/admin/innovation/delete/<int:inn_id>')
@login_required
def delete_innovation(inn_id):
    conn = get_db_connection()
    inn = conn.execute('SELECT * FROM innovations WHERE id = ?', (inn_id,)).fetchone()
    if not inn:
        flash('Inovasi tidak ditemukan!', 'warning')
        conn.close()
        return redirect(url_for('innovations'))
    if session.get('role') == 'admin_pikr' or (session.get('role') == 'konselor' and inn['author'] == session.get('username')):
        conn.execute('DELETE FROM innovations WHERE id = ?', (inn_id,))
        conn.commit()
        notify_admins(f"🗑️ Inovasi '{inn['title']}' dihapus oleh {session['username']}", url_for('innovations'))
        flash('Inovasi berhasil dihapus!', 'success')
    else:
        flash('Akses ditolak!', 'danger')
    conn.close()
    return redirect(url_for('innovations'))

# Add comment to innovation
@app.route('/inovasi/add_comment/<int:inn_id>', methods=['POST'])
@login_required
def add_innovation_comment(inn_id):
    comment_text = request.form.get('comment')
    rating = request.form.get('rating')
    conn = get_db_connection()
    conn.execute('INSERT INTO innovations_comments (innovation_id, username, text, rating) VALUES (?, ?, ?, ?)',
                 (inn_id, session['username'], comment_text, int(rating)))
    conn.commit()
    add_points(session['username'], 10)
    # Notification to author
    article = conn.execute('SELECT * FROM innovations WHERE id = ?', (inn_id,)).fetchone()
    if article and article['author'] != session.get('username'):
        add_notification(article['author'], f"💬 Ada ulasan baru di inovasi '{article['title']}'", url_for('innovations'))
    notify_admins(f"💬 Ulasan baru dari {session['username']} di inovasi.", url_for('innovations'))
    conn.close()
    flash('Ulasan berhasil ditambahkan!', 'success')
    return redirect(url_for('innovations'))

# Edit comment for innovation
@app.route('/inovasi/edit_comment/<int:inn_id>/<int:comment_id>', methods=['POST'])
@login_required
def edit_innovation_comment(inn_id, comment_id):
    new_text = request.form.get('comment')
    conn = get_db_connection()
    c = conn.execute('SELECT * FROM innovations_comments WHERE id = ? AND innovation_id = ?', (comment_id, inn_id)).fetchone()
    if c:
        if c['username'] == session.get('username'):
            conn.execute('UPDATE innovations_comments SET text = ? WHERE id = ?', (new_text, comment_id))
            conn.commit()
            flash('Ulasan berhasil diperbarui!', 'success')
        else:
            flash('Akses ditolak!', 'danger')
    conn.close()
    return redirect(url_for('innovation_detail', inn_id=inn_id))

# Delete comment for innovation
@app.route('/admin/innovation/delete_comment/<int:inn_id>/<int:comment_id>')
@login_required
def delete_innovation_comment(inn_id, comment_id):
    conn = get_db_connection()
    c = conn.execute('SELECT * FROM innovations_comments WHERE id = ? AND innovation_id = ?', (comment_id, inn_id)).fetchone()
    if c:
        if session.get('role') == 'admin_pikr' or c['username'] == session.get('username'):
            conn.execute('DELETE FROM innovations_comments WHERE id = ?', (comment_id,))
            conn.commit()
            flash('Ulasan berhasil dihapus!', 'info')
        else:
            flash('Akses ditolak!', 'danger')
    conn.close()
    return redirect(url_for('innovation_detail', inn_id=inn_id))

@app.route('/edukasi/<int:article_id>')
def education_detail(article_id):
    conn = get_db_connection()
    art = conn.execute('SELECT * FROM education WHERE id = ?', (article_id,)).fetchone()
    if art:
        article = dict(art)
        article['comments'] = [dict(c) for c in conn.execute('SELECT * FROM education_comments WHERE article_id = ?', (article_id,)).fetchall()]
        conn.close()
        return render_template('education_detail.html', article=article)
    conn.close()
    flash("Artikel tidak ditemukan", "warning")
    return redirect(url_for('education_event', tab='edukasi'))

@app.route('/upload_image', methods=['POST'])
@login_required
def upload_image():
    file_gambar = request.files.get('upload')
    if file_gambar:
        nama_gambar = secure_filename(file_gambar.filename)
        nama_gambar = f"{int(datetime.now().timestamp())}_{nama_gambar}"
        file_gambar.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_gambar))
        
        url_gambar = url_for('static', filename=f'uploads/{nama_gambar}')
        return f"""
        <script type='text/javascript'>
            window.parent.CKEDITOR.tools.callFunction({request.args.get('CKEditorFuncNum')}, '{url_gambar}', 'Gambar berhasil disisipkan!');
        </script>
        """
    return ''

@app.route('/add_article', methods=['POST'])
@login_required
def add_article():
    if session.get('role') not in ['konselor', 'admin_pikr']:
        flash("Akses ditolak!", "danger")
        return redirect(url_for('education_event', tab='edukasi'))
    
    title = request.form.get('title')
    content = request.form.get('content', '')
    
    file_dokumen = request.files.get('dokumen')
    nama_dokumen = None
    
    if file_dokumen and file_dokumen.filename != '':
        nama_dokumen = secure_filename(file_dokumen.filename)
        nama_dokumen = f"{int(datetime.now().timestamp())}_{nama_dokumen}"
        file_dokumen.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen))

    if (not content or content.strip() == '') and not nama_dokumen:
        flash("Gagal menerbitkan: Mohon isi teks materi atau pilih file dokumen PDF/DOCX untuk diupload!", "danger")
        return redirect(url_for('education_event', tab='edukasi'))

    conn = get_db_connection()
    conn.execute('INSERT INTO education (title, content, author, author_id, dokumen) VALUES (?, ?, ?, ?, ?)',
                 (title, content, session['username'], session['user_id'], nama_dokumen))
    conn.commit()
    conn.close()
    
    notify_role('anggota_remaja', f"📚 Materi Baru: {title}", url_for('education_event', tab='edukasi'))
    notify_admins(f"📚 Materi Baru diterbitkan oleh {session['username']}: {title}", url_for('education_event', tab='edukasi'))
    
    flash("Materi edukasi berhasil diterbitkan!", "success")
    return redirect(url_for('education_event', tab='edukasi'))

@app.route('/admin/education/add', methods=['POST'])
@login_required
def add_education_admin():
    return add_article()

@app.route('/admin/education/delete/<int:edu_id>')
@login_required
def delete_education(edu_id):
    conn = get_db_connection()
    article = conn.execute('SELECT * FROM education WHERE id = ?', (edu_id,)).fetchone()
    if not article:
        flash('Materi tidak ditemukan!', 'warning')
        conn.close()
        return redirect(url_for('education_event', tab='edukasi'))
        
    if session.get('role') == 'admin_pikr' or (session.get('role') == 'konselor' and article['author'] == session.get('username')):
        conn.execute('DELETE FROM education WHERE id = ?', (edu_id,))
        conn.commit()
        notify_admins(f"🗑️ Artikel '{article['title']}' dihapus oleh {session['username']}", url_for('education_event', tab='edukasi'))
        flash('Materi edukasi berhasil dihapus!', 'success')
    else:
        flash('Akses ditolak: Anda tidak berhak menghapus materi ini!', 'danger')
    conn.close()
    return redirect(url_for('education_event', tab='edukasi'))

@app.route('/add_comment/<int:article_id>', methods=['POST'])
@login_required
def add_comment(article_id):
    comment_text = request.form.get('comment')
    rating = request.form.get('rating')

    conn = get_db_connection()
    conn.execute('INSERT INTO education_comments (article_id, username, text, rating) VALUES (?, ?, ?, ?)',
                 (article_id, session['username'], comment_text, int(rating)))
    conn.commit()
    
    add_points(session['username'], 10)
    
    article = conn.execute('SELECT * FROM education WHERE id = ?', (article_id,)).fetchone()
    if article and article['author'] != session.get('username'):
        add_notification(article['author'], f"💬 Ada ulasan baru di artikel '{article['title']}'", url_for('education_event', tab='edukasi'))
    if article:
        notify_admins(f"💬 Ulasan baru dari {session['username']} di artikel: {article['title']}", url_for('education_event', tab='edukasi'))
    conn.close()
        
    flash('Ulasan berhasil ditambahkan!', 'success')
    return redirect(url_for('education_event', tab='edukasi'))

@app.route('/edit_comment/<int:article_id>/<int:comment_id>', methods=['POST'])
@login_required
def edit_comment(article_id, comment_id):
    new_text = request.form.get('comment')
    conn = get_db_connection()
    c = conn.execute('SELECT * FROM education_comments WHERE id = ? AND article_id = ?', (comment_id, article_id)).fetchone()
    if c:
        if c['username'] == session.get('username'):
            conn.execute('UPDATE education_comments SET text = ? WHERE id = ?', (new_text, comment_id))
            conn.commit()
            flash('Ulasan berhasil diperbarui!', 'success')
        else:
            flash('Akses ditolak!', 'danger')
    conn.close()
    return redirect(url_for('education_event', tab='edukasi'))

@app.route('/admin/education/delete_comment/<int:article_id>/<int:comment_id>')
@login_required
def delete_comment(article_id, comment_id):
    conn = get_db_connection()
    c = conn.execute('SELECT * FROM education_comments WHERE id = ? AND article_id = ?', (comment_id, article_id)).fetchone()
    if c:
        if session.get('role') == 'admin_pikr' or c['username'] == session.get('username'):
            conn.execute('DELETE FROM education_comments WHERE id = ?', (comment_id,))
            conn.commit()
            flash('Ulasan berhasil dihapus!', 'info')
        else:
            flash('Akses ditolak!', 'danger')
    conn.close()
    return redirect(url_for('education_event', tab='edukasi'))

@app.route('/admin/education/edit/<int:edu_id>', methods=['GET', 'POST'])
@login_required
def edit_education(edu_id):
    conn = get_db_connection()
    article = conn.execute('SELECT * FROM education WHERE id = ?', (edu_id,)).fetchone()
    if not article:
        conn.close()
        if request.method == 'GET':
            return jsonify({'error': 'Materi tidak ditemukan'}), 404
        flash('Materi tidak ditemukan!', 'warning')
        return redirect(url_for('education_event', tab='edukasi'))

    # Permission check
    if session.get('role') not in ['admin_pikr', 'konselor'] or \
       (session.get('role') == 'konselor' and article['author'] != session.get('username')):
        conn.close()
        if request.method == 'GET':
            return jsonify({'error': 'Akses ditolak'}), 403
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('education_event', tab='edukasi'))

    if request.method == 'GET':
        data = dict(article)
        conn.close()
        return jsonify(data)

    # POST: save edits
    title = request.form.get('title', article['title'])
    content = request.form.get('content', article['content'])

    file_dokumen = request.files.get('dokumen')
    nama_dokumen = article['dokumen']
    if file_dokumen and file_dokumen.filename != '':
        nama_dokumen = secure_filename(file_dokumen.filename)
        nama_dokumen = f"{int(datetime.now().timestamp())}_{nama_dokumen}"
        file_dokumen.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen))

    conn.execute('UPDATE education SET title = ?, content = ?, dokumen = ? WHERE id = ?',
                 (title, content, nama_dokumen, edu_id))
    conn.commit()
    conn.close()
    flash('Materi edukasi berhasil diperbarui!', 'success')
    return redirect(url_for('education_event', tab='edukasi'))

# --- CHAT PRIVAT FITUR ---
def send_message(sender, receiver, message_text, attachment=None, is_bot=False):
    conn = get_db_connection()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    conn.execute('INSERT INTO messages (sender, receiver, message, attachment, timestamp, seen, is_bot) VALUES (?, ?, ?, ?, ?, ?, ?)',
                 (sender, receiver, message_text, attachment, timestamp, 0, 1 if is_bot else 0))
    conn.commit()
    conn.close()
    
    if not is_bot:
        add_notification(receiver, f"💬 Pesan baru dari {sender}", url_for('chat', receiver_username=sender))

    # --- EMERGENCY DETECTION SYSTEM ---
    msg_lower = message_text.lower()
    
    # Kosa Kata Darurat (Merah)
    EMERGENCY_KEYWORDS = [
        "bunuh diri", "akhiri hidup", "pengen mati", "mati aja", "sayat", "potong urat", 
        "racun", "gantung diri", "loncat", "darah", "sekarat", "tolong cepat",
        "suicide", "kill myself", "end my life", "self harm"
    ]
    
    # Kosa Kata Prioritas (Kuning)
    PRIORITY_KEYWORDS = [
        "pembullyan", "bully", "dihina", "diejek", "dikucilkan", "diteror", "intimidasi", 
        "pelecehan", "depresi", "stress", "tekanan", "trauma", "skizofrenia", "bipolar", 
        "halusinasi", "delusi", "gangguan jiwa", "skizo", "odgj", "gangguan mental", 
        "serangan panik", "bullying", "harassment", "abused", "depressed", 
        "schizophrenia", "panic attack", "borderline", "bpd", "ocd"
    ]

    level = None
    if any(k in msg_lower for k in EMERGENCY_KEYWORDS):
        level = "emergency"
    elif any(k in msg_lower for k in PRIORITY_KEYWORDS):
        level = "high"

    if level:
        conn2 = get_db_connection()
        # Temukan sesi antara dua user ini, jangan turunkan dari emergency ke high
        sessions_found = conn2.execute('''
            SELECT * FROM sessions 
            WHERE (member_name = ? AND counselor_name = ?) OR (member_name = ? AND counselor_name = ?)
        ''', (sender, receiver, receiver, sender)).fetchall()
        
        for s in sessions_found:
            if s['priority'] == 'emergency' and level == 'high':
                continue
            conn2.execute('UPDATE sessions SET priority = ? WHERE id = ?', (level, s['id']))
        conn2.commit()
        conn2.close()
        
        if sessions_found:
            if level == 'emergency':
                notify_role('klinik_kesehatan', f"🚨 DARURAT: Deteksi kosa kata kritis dari {sender}!", url_for('counseling'))
                notify_admins(f"🚨 DARURAT: {sender} membutuhkan bantuan segera!", url_for('counseling'))
            else:
                notify_role('klinik_kesehatan', f"⚠️ PRIORITAS: Deteksi topik berisiko (Misal: bully) dari {sender}!", url_for('counseling'))
                notify_admins(f"⚠️ PRIORITAS: {sender} membahas topik sensitif/berisiko.", url_for('counseling'))
                add_notification(receiver, f"⚠️ PRIORITAS: Pesan dari {sender} terdeteksi butuh perhatian khusus.", url_for('counseling'))

@app.route('/delete_message/<int:message_id>/<string:receiver_username>')
@login_required
def delete_message(message_id, receiver_username):
    conn = get_db_connection()
    m = conn.execute('SELECT * FROM messages WHERE id = ?', (message_id,)).fetchone()
    if m and (m['sender'] == session['username'] or session.get('role') == 'admin_pikr'):
        conn.execute('DELETE FROM messages WHERE id = ?', (message_id,))
        conn.commit()
        flash("Pesan dihapus", "info")
    else:
        flash("Gagal menghapus pesan", "danger")
    conn.close()
    return redirect(url_for('chat', receiver_username=receiver_username))

@app.route('/chat')
@app.route('/chat/<string:receiver_username>', methods=['GET', 'POST'])
@login_required
def chat(receiver_username=None):
    my_username = session.get('username')
    
    if receiver_username and my_username == receiver_username:
        flash("Anda tidak bisa mengirim pesan ke diri sendiri.", "warning")
        return redirect(url_for('chat'))

    if request.method == 'POST' and receiver_username:
        message_text = request.form.get('message', '')
        attachment_path = None
        
        file = request.files.get('attachment')
        if file and file.filename:
            filename = f"chat_{int(datetime.now().timestamp())}_{secure_filename(file.filename)}"
            filepath = os.path.join('static/uploads/chats', filename)
            os.makedirs('static/uploads/chats', exist_ok=True)
            file.save(filepath)
            attachment_path = filepath
            
        if message_text.strip() or attachment_path:
            send_message(my_username, receiver_username, message_text, attachment_path)
            
        return redirect(url_for('chat', receiver_username=receiver_username))
        
    conn = get_db_connection()
    chat_partners = [row['partner'] for row in conn.execute('''
        SELECT receiver as partner FROM messages WHERE sender = ?
        UNION
        SELECT sender as partner FROM messages WHERE receiver = ?
    ''', (my_username, my_username)).fetchall()]
    
    filtered_messages = []
    if receiver_username:
        filtered_messages = [dict(m) for m in conn.execute('''
            SELECT * FROM messages 
            WHERE (sender = ? AND receiver = ?) OR (sender = ? AND receiver = ?)
            ORDER BY id ASC
        ''', (my_username, receiver_username, receiver_username, my_username)).fetchall()]
    conn.close()
        
    return render_template('chat.html', 
                           receiver_username=receiver_username, 
                           history=filtered_messages, 
                           chat_partners=sorted(chat_partners))

@app.route('/clear_chat/<string:receiver_username>', methods=['POST'])
@login_required
def clear_chat(receiver_username):
    current_user = session.get('username')
    conn = get_db_connection()
    conn.execute('''
        DELETE FROM messages 
        WHERE (sender = ? AND receiver = ?) OR (sender = ? AND receiver = ?)
    ''', (current_user, receiver_username, receiver_username, current_user))
    conn.commit()
    conn.close()
    flash(f"Riwayat chat dengan {receiver_username} telah dibersihkan.", "info")
    return redirect(url_for('chat', receiver_username=receiver_username))

@app.route('/transfer_to_clinic/<string:receiver_username>', methods=['POST'])
@login_required
def transfer_to_clinic(receiver_username):
    if session.get('role') != 'konselor':
        flash("Akses ditolak.", "danger")
        return redirect(url_for('chat'))
    
    sender = session.get('username')
    
    # Notify member to chat with clinic
    add_notification(receiver_username, f"🏥 Konselor {sender} merujukmu ke Klinik Kesehatan. Klik di sini untuk mulai obrolan.", url_for('chat', receiver_username='klinik'))
    
    # Notify clinic that a chat has been transferred
    notify_role('klinik_kesehatan', f"🏥 Konselor {sender} merujuk {receiver_username} ke Klinik. Segera tindak lanjuti.", url_for('chat', receiver_username=receiver_username))
    
    # Send bot message to the current chat
    bot_msg = f"Sesi ini telah dirujuk ke Klinik Kesehatan oleh Konselor. {receiver_username}, silakan hubungi akun 'klinik' untuk penanganan medis lebih lanjut."
    send_message(sender, receiver_username, bot_msg, is_bot=True)
    
    flash("Sesi berhasil dialihkan ke Klinik Kesehatan.", "success")
    return redirect(url_for('chat', receiver_username=receiver_username))

# --- EVENTS FITUR ---
@app.route('/events')
def events():
    conn = get_db_connection()
    events_raw = conn.execute('SELECT * FROM events ORDER BY date DESC').fetchall()
    all_events = []
    for e in events_raw:
        ev = dict(e)
        participants = conn.execute('SELECT username FROM event_participants WHERE event_id = ?', (e['id'],)).fetchall()
        ev['participants'] = [p['username'] for p in participants]
        if ev.get('foto'):
            ev['foto_list'] = [f.strip() for f in ev['foto'].split(',') if f.strip()]
        else:
            ev['foto_list'] = []
        all_events.append(ev)
    conn.close()
    return render_template('events.html', events=all_events)

@app.route('/admin/events/add', methods=['POST'])
@login_required
def add_event():
    if session.get('role') not in ['admin_pikr', 'konselor']:
        return redirect(url_for('events'))
    
    title = request.form.get('title')
    description = request.form.get('description')
    date = request.form.get('date')
    time = request.form.get('time')

    # Handle multiple foto uploads
    uploaded_photos = []
    files = request.files.getlist('foto')
    for file_foto in files:
        if file_foto and file_foto.filename != '':
            nama_foto = secure_filename(file_foto.filename)
            nama_foto = f"{int(datetime.now().timestamp())}_{nama_foto}"
            file_foto.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_foto))
            uploaded_photos.append(nama_foto)
    
    foto_str = ",".join(uploaded_photos) if uploaded_photos else None

    conn = get_db_connection()
    conn.execute('INSERT INTO events (title, description, date, time, author, foto) VALUES (?, ?, ?, ?, ?, ?)',
                 (title, description, date, time, session.get('username'), foto_str))
    conn.commit()
    conn.close()
    
    notify_role('anggota_remaja', f"🎉 Event Baru: {title}", url_for('events'))
    notify_admins(f"🎉 Event Baru ditambahkan oleh {session['username']}: {title}", url_for('events'))
    
    flash('Kegiatan berhasil ditambahkan!', 'success')
    return redirect(url_for('events'))

@app.route('/admin/events/delete/<int:event_id>')
@login_required
def delete_event(event_id):
    if session.get('role') not in ['admin_pikr', 'konselor']:
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('events'))
    conn = get_db_connection()
    ev = conn.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    if not ev:
        conn.close()
        flash('Kegiatan tidak ditemukan!', 'warning')
        return redirect(url_for('events'))
    # Konselor hanya bisa hapus event miliknya sendiri
    if session.get('role') == 'konselor' and ev['author'] != session.get('username'):
        conn.close()
        flash('Akses ditolak: Anda hanya bisa menghapus kegiatan milik Anda sendiri!', 'danger')
        return redirect(url_for('events'))
    if ev['foto']:
        photos = [f.strip() for f in ev['foto'].split(',') if f.strip()]
        for photo in photos:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], photo)
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                except Exception as e:
                    print("Error deleting event photo:", e)
    conn.execute('DELETE FROM events WHERE id = ?', (event_id,))
    conn.commit()
    conn.close()
    flash('Kegiatan berhasil dihapus!', 'success')
    return redirect(url_for('events'))

@app.route('/admin/events/edit/<int:event_id>', methods=['POST'])
@login_required
def edit_event(event_id):
    if session.get('role') not in ['admin_pikr', 'konselor']:
        return redirect(url_for('events'))
    conn = get_db_connection()
    ev = conn.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    if not ev:
        conn.close()
        flash('Kegiatan tidak ditemukan!', 'warning')
        return redirect(url_for('events'))

    # Konselor hanya bisa mengedit kegiatan milik mereka sendiri
    if session.get('role') == 'konselor' and ev['author'] != session.get('username'):
        conn.close()
        flash('Akses ditolak: Anda hanya bisa mengedit kegiatan milik Anda sendiri!', 'danger')
        return redirect(url_for('events'))

    nama_foto = ev['foto']
    current_photos = [f.strip() for f in nama_foto.split(',') if f.strip()] if nama_foto else []

    # Handle deleted photos
    deleted_str = request.form.get('deleted_photos', '')
    deleted_photos = [f.strip() for f in deleted_str.split(',') if f.strip()]
    
    remaining_photos = []
    for photo in current_photos:
        if photo in deleted_photos:
            # Physically delete file
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], photo)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception as e:
                    print(f"Error deleting file {photo}: {e}")
        else:
            remaining_photos.append(photo)

    # Handle new uploaded files
    uploaded_photos = []
    files = request.files.getlist('foto')
    for file_foto in files:
        if file_foto and file_foto.filename != '':
            name_f = secure_filename(file_foto.filename)
            name_f = f"{int(datetime.now().timestamp())}_{name_f}"
            file_foto.save(os.path.join(app.config['UPLOAD_FOLDER'], name_f))
            uploaded_photos.append(name_f)
    
    final_photos = remaining_photos + uploaded_photos
    final_foto_str = ",".join(final_photos) if final_photos else None

    conn.execute('UPDATE events SET title = ?, description = ?, date = ?, time = ?, foto = ? WHERE id = ?',
                 (request.form.get('title'), request.form.get('description'),
                  request.form.get('date'), request.form.get('time'), final_foto_str, event_id))
    conn.commit()
    conn.close()
    flash('Kegiatan berhasil diperbarui!', 'success')
    return redirect(url_for('events'))

@app.route('/join_event/<int:event_id>')
@login_required
def join_event(event_id):
    if session.get('role') != 'anggota_remaja':
        return redirect(url_for('events'))
    conn = get_db_connection()
    existing = conn.execute('SELECT 1 FROM event_participants WHERE event_id = ? AND username = ?', (event_id, session['username'])).fetchone()
    if not existing:
        conn.execute('INSERT INTO event_participants (event_id, username) VALUES (?, ?)', (event_id, session['username']))
        conn.commit()
        add_points(session['username'], 50)
    
    event = conn.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    conn.close()
    if event:
        msg = f"👥 {session['username']} mendaftar ke event: {event['title']}"
        if event['author']:
            add_notification(event['author'], msg, url_for('events'))
        notify_admins(msg, url_for('events'))
        
    flash('Berhasil mendaftar ke kegiatan!', 'success')
    ref = request.referrer or ''
    if f'/event/{event_id}' in ref:
        return redirect(url_for('event_detail', event_id=event_id))
    return redirect(url_for('events'))

@app.route('/leave_event/<int:event_id>')
@login_required
def leave_event(event_id):
    if session.get('role') != 'anggota_remaja':
        return redirect(url_for('events'))
    conn = get_db_connection()
    conn.execute('DELETE FROM event_participants WHERE event_id = ? AND username = ?', (event_id, session['username']))
    conn.commit()
    conn.close()
    flash('Pendaftaran kegiatan berhasil dibatalkan.', 'info')
    # Redirect back to detail page if came from there
    ref = request.referrer or ''
    if f'/event/{event_id}' in ref:
        return redirect(url_for('event_detail', event_id=event_id))
    return redirect(url_for('events'))

# --- API: Real-time participant count ---
@app.route('/api/event/<int:event_id>/participants')
def api_event_participants(event_id):
    conn = get_db_connection()
    rows = conn.execute('SELECT username FROM event_participants WHERE event_id = ?', (event_id,)).fetchall()
    conn.close()
    participants = [r['username'] for r in rows]
    return jsonify({'count': len(participants), 'participants': participants})


# --- TENTANG KAMI FITUR ---
@app.route('/tentang-kami')
def about():
    conn = get_db_connection()
    history_row = conn.execute('SELECT content FROM about_history ORDER BY id ASC LIMIT 1').fetchone()
    history = history_row['content'] if history_row else ""
    officers = [dict(o) for o in conn.execute('SELECT * FROM about_officers ORDER BY id ASC').fetchall()]
    stakeholders = [dict(s) for s in conn.execute('SELECT * FROM about_stakeholders ORDER BY id ASC').fetchall()]
    conn.close()
    return render_template('about.html', history=history, officers=officers, stakeholders=stakeholders)

@app.route('/admin/tentang-kami/add-stakeholder', methods=['POST'])
@login_required
def add_stakeholder():
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak!", "danger")
        return redirect(url_for('about'))
    
    name = request.form.get('name')
    role_title = request.form.get('role_title')
    role_desc = request.form.get('role_desc', '')
    category = request.form.get('category', '')
    icon_type = request.form.get('icon_type', 'community')
    
    if not name or not role_title:
        flash("Nama dan peran stakeholder wajib diisi!", "danger")
        return redirect(url_for('about'))
        
    # Handle Logo Upload
    logo_filename = None
    file = request.files.get('logo')
    if file and file.filename != '':
        os.makedirs('static/uploads/stakeholders', exist_ok=True)
        logo_filename = f"logo_{int(time.time())}_{secure_filename(file.filename)}"
        file.save(os.path.join('static/uploads/stakeholders', logo_filename))
    
    conn = get_db_connection()
    conn.execute('INSERT INTO about_stakeholders (name, role_title, role_desc, category, icon_type, logo) VALUES (?, ?, ?, ?, ?, ?)',
                 (name, role_title, role_desc, category, icon_type, logo_filename))
    conn.commit()
    conn.close()
    flash("Stakeholder berhasil ditambahkan!", "success")
    return redirect(url_for('about'))

@app.route('/admin/tentang-kami/edit-stakeholder/<int:stk_id>', methods=['POST'])
@login_required
def edit_stakeholder(stk_id):
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak!", "danger")
        return redirect(url_for('about'))
    
    name = request.form.get('name')
    role_title = request.form.get('role_title')
    role_desc = request.form.get('role_desc', '')
    category = request.form.get('category', '')
    icon_type = request.form.get('icon_type', 'community')
    
    if not name or not role_title:
        flash("Nama dan peran stakeholder wajib diisi!", "danger")
        return redirect(url_for('about'))
    
    conn = get_db_connection()
    current_stk = conn.execute('SELECT * FROM about_stakeholders WHERE id = ?', (stk_id,)).fetchone()
    if not current_stk:
        conn.close()
        flash("Stakeholder tidak ditemukan!", "danger")
        return redirect(url_for('about'))
        
    logo_filename = current_stk['logo']
    
    # Handle Logo Upload Update
    file = request.files.get('logo')
    if file and file.filename != '':
        os.makedirs('static/uploads/stakeholders', exist_ok=True)
        # Delete old logo if exists
        if logo_filename:
            old_path = os.path.join('static/uploads/stakeholders', logo_filename)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except Exception:
                    pass
        logo_filename = f"logo_{int(time.time())}_{secure_filename(file.filename)}"
        file.save(os.path.join('static/uploads/stakeholders', logo_filename))
    
    conn.execute('UPDATE about_stakeholders SET name = ?, role_title = ?, role_desc = ?, category = ?, icon_type = ?, logo = ? WHERE id = ?',
                 (name, role_title, role_desc, category, icon_type, logo_filename, stk_id))
    conn.commit()
    conn.close()
    flash("Stakeholder berhasil diperbarui!", "success")
    return redirect(url_for('about'))

@app.route('/admin/tentang-kami/delete-stakeholder/<int:stk_id>')
@login_required
def delete_stakeholder(stk_id):
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak!", "danger")
        return redirect(url_for('about'))
    
    conn = get_db_connection()
    current_stk = conn.execute('SELECT * FROM about_stakeholders WHERE id = ?', (stk_id,)).fetchone()
    if current_stk:
        # Delete logo file
        logo_filename = current_stk['logo']
        if logo_filename:
            logo_path = os.path.join('static/uploads/stakeholders', logo_filename)
            if os.path.exists(logo_path):
                try:
                    os.remove(logo_path)
                except Exception:
                    pass
        conn.execute('DELETE FROM about_stakeholders WHERE id = ?', (stk_id,))
        conn.commit()
    conn.close()
    flash("Stakeholder berhasil dihapus!", "success")
    return redirect(url_for('about'))

@app.route('/admin/tentang-kami/edit-sejarah', methods=['POST'])
@login_required
def edit_about_history():
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak!", "danger")
        return redirect(url_for('about'))
    
    new_content = request.form.get('content')
    if new_content and new_content.strip():
        conn = get_db_connection()
        row = conn.execute('SELECT id FROM about_history ORDER BY id ASC LIMIT 1').fetchone()
        if row:
            conn.execute('UPDATE about_history SET content = ? WHERE id = ?', (new_content, row['id']))
        else:
            conn.execute('INSERT INTO about_history (content) VALUES (?)', (new_content,))
        conn.commit()
        conn.close()
        flash("Sejarah berhasil diperbarui!", "success")
    else:
        flash("Konten sejarah tidak boleh kosong!", "danger")
    return redirect(url_for('about'))
@app.route('/admin/tentang-kami/add-pengurus', methods=['POST'])
@login_required
def add_officer():
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak!", "danger")
        return redirect(url_for('about'))
    
    caption = request.form.get('caption', '')
    file = request.files.get('photo')
    
    photo_path = ''
    if file and file.filename:
        filename = f"group_photo_{int(time.time())}_{secure_filename(file.filename)}"
        filepath = os.path.join('static/uploads/about', filename).replace('\\', '/')
        os.makedirs('static/uploads/about', exist_ok=True)
        file.save(filepath)
        photo_path = filepath
    else:
        flash("Foto bersama wajib diunggah!", "danger")
        return redirect(url_for('about'))
    
    conn = get_db_connection()
    # Name dan position dihapus dari query INSERT
    conn.execute('INSERT INTO about_officers (photo, caption) VALUES (?, ?)',
                 (photo_path, caption))
    conn.commit()
    conn.close()
    flash("Foto bersama berhasil ditambahkan!", "success")
    return redirect(url_for('about'))

@app.route('/admin/tentang-kami/edit-pengurus/<int:officer_id>', methods=['POST'])
@login_required
def edit_officer(officer_id):
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak!", "danger")
        return redirect(url_for('about'))
    
    caption = request.form.get('caption', '')
    file = request.files.get('photo')
    
    conn = get_db_connection()
    officer = conn.execute('SELECT * FROM about_officers WHERE id = ?', (officer_id,)).fetchone()
    if not officer:
        conn.close()
        flash("Data foto tidak ditemukan!", "danger")
        return redirect(url_for('about'))
    
    photo_path = officer['photo']
    if file and file.filename:
        if officer['photo'] and os.path.exists(officer['photo']):
            try:
                os.remove(officer['photo'])
            except Exception as e:
                print("Error deleting old photo:", e)
        
        filename = f"group_photo_{int(time.time())}_{secure_filename(file.filename)}"
        filepath = os.path.join('static/uploads/about', filename).replace('\\', '/')
        os.makedirs('static/uploads/about', exist_ok=True)
        file.save(filepath)
        photo_path = filepath
    
    # Name dan position dihapus dari query UPDATE
    conn.execute('UPDATE about_officers SET photo = ?, caption = ? WHERE id = ?',
                 (photo_path, caption, officer_id))
    conn.commit()
    conn.close()
    flash("Foto bersama berhasil diperbarui!", "success")
    return redirect(url_for('about'))

@app.route('/admin/tentang-kami/delete-pengurus/<int:officer_id>', methods=['GET', 'POST'])
@login_required
def delete_officer(officer_id):
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak!", "danger")
        return redirect(url_for('about'))
    
    conn = get_db_connection()
    officer = conn.execute('SELECT * FROM about_officers WHERE id = ?', (officer_id,)).fetchone()
    if officer:
        if officer['photo'] and os.path.exists(officer['photo']):
            try:
                os.remove(officer['photo'])
            except Exception as e:
                print("Error deleting photo file:", e)
        conn.execute('DELETE FROM about_officers WHERE id = ?', (officer_id,))
        conn.commit()
        flash("Data foto berhasil dihapus!", "success")
    else:
        flash("Data foto tidak ditemukan!", "danger")
    conn.close()
    return redirect(url_for('about'))

# --- PRESTASI FITUR ---
@app.route('/prestasi')
def achievements():
    conn = get_db_connection()
    ach_raw = conn.execute('SELECT * FROM achievements ORDER BY id DESC').fetchall()
    achievements_list = []
    for ach in ach_raw:
        a = dict(ach)
        a['comments'] = [dict(c) for c in conn.execute('SELECT * FROM achievements_comments WHERE achievement_id = ?', (a['id'],)).fetchall()]
        # Parse dokumen as foto_list
        if a.get('dokumen'):
            a['foto_list'] = [f.strip() for f in a['dokumen'].split(',') if f.strip()]
        else:
            a['foto_list'] = []
        achievements_list.append(a)
    conn.close()
    return render_template('achievements.html', achievements=achievements_list)

@app.route('/prestasi/<int:achievement_id>')
def achievement_detail(achievement_id):
    conn = get_db_connection()
    ach = conn.execute('SELECT * FROM achievements WHERE id = ?', (achievement_id,)).fetchone()
    if ach:
        achievement = dict(ach)
        achievement['comments'] = [dict(c) for c in conn.execute('SELECT * FROM achievements_comments WHERE achievement_id = ?', (achievement_id,)).fetchall()]
        # Parse dokumen as foto_list
        if achievement.get('dokumen'):
            achievement['foto_list'] = [f.strip() for f in achievement['dokumen'].split(',') if f.strip()]
        else:
            achievement['foto_list'] = []
        conn.close()
        return render_template('achievement_detail.html', achievement=achievement)
    conn.close()
    flash("Prestasi tidak ditemukan", "warning")
    return redirect(url_for('achievements'))

@app.route('/admin/prestasi/add', methods=['POST'])
@login_required
def add_achievement():
    if session.get('role') not in ['konselor', 'admin_pikr']:
        flash("Akses ditolak!", "danger")
        return redirect(url_for('achievements'))
    
    title = request.form.get('title')
    content = request.form.get('content', '')
    
    # Handle multiple file uploads
    uploaded_photos = []
    files = request.files.getlist('dokumen')
    for file_f in files:
        if file_f and file_f.filename != '':
            nama = secure_filename(file_f.filename)
            nama = f"{int(datetime.now().timestamp())}_{nama}"
            file_f.save(os.path.join(app.config['UPLOAD_FOLDER'], nama))
            uploaded_photos.append(nama)
    
    nama_dokumen = ','.join(uploaded_photos) if uploaded_photos else None

    if (not content or content.strip() == '') and not nama_dokumen:
        flash("Gagal menerbitkan: Mohon isi teks prestasi atau pilih file dokumen/foto untuk diupload!", "danger")
        return redirect(url_for('achievements'))

    date_input = request.form.get('date')
    if date_input:
        date_str = date_input.replace('T', ' ')
    else:
        date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    conn = get_db_connection()
    conn.execute('INSERT INTO achievements (title, content, author, author_id, dokumen, date) VALUES (?, ?, ?, ?, ?, ?)',
                 (title, content, session['username'], session['user_id'], nama_dokumen, date_str))
    conn.commit()
    conn.close()
    
    notify_role('anggota_remaja', f"🏆 Prestasi Baru: {title}", url_for('achievements'))
    notify_admins(f"🏆 Prestasi Baru diterbitkan oleh {session['username']}: {title}", url_for('achievements'))
    
    flash("Prestasi berhasil diterbitkan!", "success")
    return redirect(url_for('achievements'))

@app.route('/admin/prestasi/delete/<int:ach_id>')
@login_required
def delete_achievement(ach_id):
    conn = get_db_connection()
    achievement = conn.execute('SELECT * FROM achievements WHERE id = ?', (ach_id,)).fetchone()
    if not achievement:
        flash('Prestasi tidak ditemukan!', 'warning')
        conn.close()
        return redirect(url_for('achievements'))
        
    if session.get('role') == 'admin_pikr' or (session.get('role') == 'konselor' and achievement['author'] == session.get('username')):
        if achievement['dokumen']:
            photos = [f.strip() for f in achievement['dokumen'].split(',') if f.strip()]
            for photo in photos:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], photo)
                if os.path.exists(filepath):
                    try:
                        os.remove(filepath)
                    except Exception as e:
                        print("Error deleting document:", e)
                    
        conn.execute('DELETE FROM achievements WHERE id = ?', (ach_id,))
        conn.commit()
        notify_admins(f"🗑️ Prestasi '{achievement['title']}' dihapus oleh {session['username']}", url_for('achievements'))
        flash('Prestasi berhasil dihapus!', 'success')
    else:
        flash('Akses ditolak: Anda tidak berhak menghapus prestasi ini!', 'danger')
    conn.close()
    return redirect(url_for('achievements'))

@app.route('/admin/prestasi/edit/<int:ach_id>', methods=['GET', 'POST'])
@login_required
def edit_achievement(ach_id):
    conn = get_db_connection()
    achievement = conn.execute('SELECT * FROM achievements WHERE id = ?', (ach_id,)).fetchone()
    if not achievement:
        conn.close()
        if request.method == 'GET':
            return jsonify({'error': 'Not found'}), 404
        flash('Prestasi tidak ditemukan!', 'warning')
        return redirect(url_for('achievements'))

    # Only admin or the author (konselor) can edit
    if not (session.get('role') == 'admin_pikr' or
            (session.get('role') == 'konselor' and achievement['author'] == session.get('username'))):
        conn.close()
        if request.method == 'GET':
            return jsonify({'error': 'Forbidden'}), 403
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('achievements'))

    if request.method == 'GET':
        conn.close()
        db_date = achievement['date'] or ''
        formatted_date = db_date.replace(' ', 'T') if db_date else ''
        return jsonify({
            'id': achievement['id'],
            'title': achievement['title'],
            'content': achievement['content'] or '',
            'dokumen': achievement['dokumen'] or '',
            'date': formatted_date
        })

    # POST – save edits
    title = request.form.get('title', '').strip()
    content = request.form.get('content', '').strip()
    date_input = request.form.get('date')
    if not title:
        flash('Judul tidak boleh kosong!', 'danger')
        conn.close()
        return redirect(url_for('achievements'))

    nama_dokumen = achievement['dokumen']
    current_photos = [f.strip() for f in nama_dokumen.split(',') if f.strip()] if nama_dokumen else []

    # Handle deleted photos
    deleted_str = request.form.get('deleted_photos', '')
    deleted_photos = [f.strip() for f in deleted_str.split(',') if f.strip()]
    remaining_photos = []
    for photo in current_photos:
        if photo in deleted_photos:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], photo)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass
        else:
            remaining_photos.append(photo)

    # Handle new uploaded files (can be multiple images)
    uploaded_photos = []
    files = request.files.getlist('dokumen')
    for file in files:
        if file and file.filename != '':
            name_f = secure_filename(file.filename)
            name_f = f"{int(datetime.now().timestamp())}_{name_f}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], name_f).replace('\\', '/'))
            uploaded_photos.append(name_f)

    final_photos = remaining_photos + uploaded_photos
    final_dokumen = ','.join(final_photos) if final_photos else None

    if date_input:
        date_str = date_input.replace('T', ' ')
    else:
        date_str = achievement['date'] or datetime.now().strftime("%Y-%m-%d %H:%M")

    conn.execute(
        'UPDATE achievements SET title = ?, content = ?, dokumen = ?, date = ? WHERE id = ?',
        (title, content, final_dokumen, date_str, ach_id)
    )
    conn.commit()
    conn.close()
    flash('Prestasi berhasil diperbarui! ✅', 'success')
    notify_admins(f"✏️ Prestasi '{title}' diperbarui oleh {session['username']}", url_for('achievements'))
    return redirect(url_for('achievements'))

@app.route('/prestasi/add_comment/<int:ach_id>', methods=['POST'])
@login_required
def add_achievement_comment(ach_id):
    comment_text = request.form.get('comment')
    rating = request.form.get('rating')

    conn = get_db_connection()
    conn.execute('INSERT INTO achievements_comments (achievement_id, username, text, rating) VALUES (?, ?, ?, ?)',
                 (ach_id, session['username'], comment_text, int(rating)))
    conn.commit()
    
    add_points(session['username'], 5)
    
    achievement = conn.execute('SELECT * FROM achievements WHERE id = ?', (ach_id,)).fetchone()
    if achievement and achievement['author'] != session.get('username'):
        add_notification(achievement['author'], f"💬 Ada ulasan baru di prestasi '{achievement['title']}'", url_for('achievements'))
    if achievement:
        notify_admins(f"💬 Ulasan baru dari {session['username']} di prestasi: {achievement['title']}", url_for('achievements'))
    conn.close()
        
    flash('Ulasan prestasi berhasil ditambahkan!', 'success')
    return redirect(url_for('achievements'))

@app.route('/prestasi/edit_comment/<int:ach_id>/<int:comment_id>', methods=['POST'])
@login_required
def edit_achievement_comment(ach_id, comment_id):
    new_text = request.form.get('comment')
    conn = get_db_connection()
    c = conn.execute('SELECT * FROM achievements_comments WHERE id = ? AND achievement_id = ?', (comment_id, ach_id)).fetchone()
    if c:
        if c['username'] == session.get('username'):
            conn.execute('UPDATE achievements_comments SET text = ? WHERE id = ?', (new_text, comment_id))
            conn.commit()
            flash('Ulasan berhasil diperbarui!', 'success')
        else:
            flash('Akses ditolak!', 'danger')
    conn.close()
    return redirect(url_for('achievements'))

@app.route('/prestasi/delete_comment/<int:ach_id>/<int:comment_id>')
@login_required
def delete_achievement_comment(ach_id, comment_id):
    conn = get_db_connection()
    c = conn.execute('SELECT * FROM achievements_comments WHERE id = ? AND achievement_id = ?', (comment_id, ach_id)).fetchone()
    if c:
        if session.get('role') == 'admin_pikr' or c['username'] == session.get('username'):
            conn.execute('DELETE FROM achievements_comments WHERE id = ?', (comment_id,))
            conn.commit()
            flash('Ulasan berhasil dihapus!', 'info')
        else:
            flash('Akses ditolak!', 'danger')
    conn.close()
    return redirect(url_for('achievements'))

# --- ADMIN KELOLA USER ---
@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
def admin_users():
    if session.get('role') != 'admin_pikr':
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    if request.method == 'POST':
        new_username = request.form.get('username')
        new_password = request.form.get('password')
        new_role = request.form.get('role')
        birth_date = request.form.get('birth_date', '2005-01-01')

        conn.execute('INSERT INTO users (username, password, birth_date, role) VALUES (?, ?, ?, ?)',
                     (new_username, new_password, birth_date, new_role))
        conn.commit()
        conn.close()
        return redirect(url_for('admin_users'))

    all_users = [dict(u) for u in conn.execute('SELECT * FROM users ORDER BY id DESC').fetchall()]
    conn.close()
    return render_template('admin_users.html', users=all_users)

@app.route('/admin/delete_user/<role>/<int:user_id>')
@login_required
def delete_user(role, user_id):
    if session.get('role') != 'admin_pikr':
        return redirect(url_for('dashboard'))
    conn = get_db_connection()
    conn.execute('DELETE FROM users WHERE id = ? AND role = ?', (user_id, role))
    conn.commit()
    conn.close()
    flash("User berhasil dihapus!", "danger")
    return redirect(url_for('admin_users'))

# --- ADMIN DATABASE EXPORT & REPORT ---
@app.route('/admin/export')
@login_required
def admin_export():
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak: Halaman ini hanya untuk Admin.", "danger")
        return redirect(url_for('dashboard'))
    return render_template('admin_export.html')

@app.route('/admin/panel')
@login_required
def admin_panel():
    """Halaman Admin Panel visual — kelola semua data dari satu tempat"""
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak: Halaman ini hanya untuk Admin PIK-R.", "danger")
        return redirect(url_for('dashboard'))
    return render_template('admin_panel.html')

@app.route('/admin/export/backup')
@login_required
def admin_export_backup():
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak: Hanya Admin yang dapat mengunduh backup database.", "danger")
        return redirect(url_for('dashboard'))
    try:
        return send_file(DB_PATH, as_attachment=True, download_name=f"pikr_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
    except Exception as e:
        flash(f"Gagal mengunduh backup database: {str(e)}", "danger")
        return redirect(url_for('admin_export'))

@app.route('/admin/export/download')
@login_required
def admin_export_download():
    if session.get('role') != 'admin_pikr':
        flash("Akses ditolak: Hanya Admin yang dapat mengunduh data.", "danger")
        return redirect(url_for('dashboard'))
    
    table = request.args.get('table')
    file_format = request.args.get('format', 'csv')
    
    if table not in ['users', 'sessions', 'events', 'forum_posts', 'innovations', 'achievements']:
        flash("Tabel tidak valid!", "danger")
        return redirect(url_for('admin_export'))
        
    conn = get_db_connection()
    
    if table == 'users':
        rows_raw = conn.execute('SELECT id, username, birth_date, role, points, bio, profile_pic FROM users ORDER BY id ASC').fetchall()
        headers = ['ID', 'Username', 'Tanggal Lahir', 'Role', 'Poin', 'Bio', 'Foto Profil']
        rows = [[r['id'], r['username'], r['birth_date'], r['role'], r['points'], r['bio'], r['profile_pic']] for r in rows_raw]
        
    elif table == 'sessions':
        rows_raw = conn.execute('SELECT id, member_name, counselor_name, topic, date, time, status, priority FROM sessions ORDER BY id DESC').fetchall()
        headers = ['ID Sesi', 'Nama Anggota', 'Nama Konselor', 'Topik', 'Tanggal Sesi', 'Waktu Sesi', 'Status Sesi', 'Prioritas']
        rows = [[r['id'], r['member_name'], r['counselor_name'], r['topic'], r['date'], r['time'], r['status'], r['priority']] for r in rows_raw]
        
    elif table == 'events':
        rows_raw = conn.execute('SELECT id, title, description, date, time, author FROM events ORDER BY id DESC').fetchall()
        headers = ['ID Event', 'Judul Event', 'Deskripsi', 'Tanggal', 'Waktu', 'Pembuat', 'Jumlah Peserta']
        rows = []
        for r in rows_raw:
            part_count = conn.execute('SELECT COUNT(*) as c FROM event_participants WHERE event_id = ?', (r['id'],)).fetchone()['c']
            rows.append([r['id'], r['title'], r['description'], r['date'], r['time'], r['author'], part_count])
            
    elif table == 'forum_posts':
        rows_raw = conn.execute('SELECT id, username, role, content, date, is_announcement FROM forum_posts ORDER BY id DESC').fetchall()
        headers = ['ID Post', 'Username', 'Role', 'Konten', 'Tanggal Kirim', 'Pengumuman (1=Ya)']
        rows = [[r['id'], r['username'], r['role'], r['content'], r['date'], r['is_announcement']] for r in rows_raw]
        
    elif table == 'innovations':
        rows_raw = conn.execute('SELECT id, title, content, author, date, dokumen FROM innovations ORDER BY id DESC').fetchall()
        headers = ['ID Inovasi', 'Judul Inovasi', 'Konten', 'Penulis', 'Tanggal', 'Lampiran Dokumen']
        rows = [[r['id'], r['title'], r['content'], r['author'], r['date'], r['dokumen']] for r in rows_raw]
        
    elif table == 'achievements':
        rows_raw = conn.execute('SELECT id, title, content, author, date, dokumen FROM achievements ORDER BY id DESC').fetchall()
        headers = ['ID Prestasi', 'Judul Prestasi', 'Konten', 'Penulis', 'Tanggal', 'Lampiran Dokumen']
        rows = [[r['id'], r['title'], r['content'], r['author'], r['date'], r['dokumen']] for r in rows_raw]
        
    conn.close()
    
    filename = f"laporan_{table}_{datetime.now().strftime('%Y%m%d')}"
    
    if file_format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = table.capitalize()[:30]
            
            header_fill = PatternFill(start_color="D8C5A3", end_color="D8C5A3", fill_type="solid")
            header_font = Font(name="Plus Jakarta Sans", size=11, bold=True, color="000000")
            
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row in rows:
                ws.append([str(val) if val is not None else '' for val in row])
                
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            ws.row_dimensions[1].height = 28
            for r in range(2, len(rows) + 2):
                ws.row_dimensions[r].height = 20
                
            ws.views.sheetView[0].showGridLines = True
            
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            
            response = make_response(out.getvalue())
            response.headers["Content-Disposition"] = f"attachment; filename={filename}.xlsx"
            response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return response
            
        except ImportError:
            flash("Modul Excel (openpyxl) tidak tersedia. Mengekspor dalam format CSV.", "warning")
            file_format = 'csv'
            
    if file_format == 'csv':
        import csv
        from io import StringIO
        
        si = StringIO()
        cw = csv.writer(si)
        cw.writerow(headers)
        cw.writerows(rows)
        
        output_data = '\ufeff' + si.getvalue()
        
        response = make_response(output_data)
        response.headers["Content-Disposition"] = f"attachment; filename={filename}.csv"
        response.headers["Content-type"] = "text/csv; charset=utf-8-sig"
        return response

def reminder_bot():
    while True:
        try:
            now = datetime.now()
            current_date = now.strftime("%Y-%m-%d")
            current_time = now.strftime("%H:%M")
            
            with app.app_context():
                conn = get_db_connection()
                sessions_due = conn.execute('''
                    SELECT * FROM sessions 
                    WHERE status = "APPROVED" AND reminder_sent = 0 AND date = ? AND time = ?
                ''', (current_date, current_time)).fetchall()
                
                for s in sessions_due:
                    member = s['member_name']
                    counselor = s['counselor_name']
                    add_notification(member, f"⏰ Mengingatkan: Sesi konselingmu dengan {counselor} dijadwalkan sekarang!", url_for('counseling'))
                    add_notification(counselor, f"⏰ Mengingatkan: Kamu ada jadwal konseling dengan {member} sekarang!", url_for('counseling'))
                    bot_msg = f"Halo! Jadwal konselingmu dengan Konselor {counselor} telah tiba. Silakan tunggu balasan dari konselor."
                    send_message(sender=counselor, receiver=member, message_text=bot_msg, is_bot=True)
                    conn.execute('UPDATE sessions SET reminder_sent = 1 WHERE id = ?', (s['id'],))
                
                conn.commit()
                conn.close()
        except Exception as e:
            print("Reminder Bot Error:", e)
        time.sleep(60)

# --- INITIALIZATION & BACKGROUND THREADS ---
# Pastikan database sudah ada (jalankan setup_db.py jika belum)
import os as _os
if not _os.path.exists(DB_PATH):
    print("[WARNING] Database belum ada! Jalankan 'python setup_db.py' terlebih dahulu.")

# Jalankan Thread Reminder Bot di Background
threading.Thread(target=reminder_bot, daemon=True).start()


# =============================================================================
# REST API v1 — Semua endpoint memerlukan header: X-API-Key: <api_key>
# Base URL: /api/v1/
# =============================================================================

# -----------------------------------------------------------------------------
# 📊 STATS — Dashboard Statistics
# -----------------------------------------------------------------------------
@app.route('/api/v1/stats', methods=['GET'])
@api_key_required
def api_stats():
    """GET /api/v1/stats — Statistik ringkas dashboard"""
    conn = get_db_connection()
    stats = {
        "total_users":        conn.execute('SELECT COUNT(*) FROM users').fetchone()[0],
        "total_konselor":     conn.execute('SELECT COUNT(*) FROM users WHERE role="konselor"').fetchone()[0],
        "total_remaja":       conn.execute('SELECT COUNT(*) FROM users WHERE role="anggota_remaja"').fetchone()[0],
        "total_admin":        conn.execute('SELECT COUNT(*) FROM users WHERE role="admin_pikr"').fetchone()[0],
        "total_sessions":     conn.execute('SELECT COUNT(*) FROM sessions').fetchone()[0],
        "pending_sessions":   conn.execute('SELECT COUNT(*) FROM sessions WHERE status="PENDING"').fetchone()[0],
        "approved_sessions":  conn.execute('SELECT COUNT(*) FROM sessions WHERE status="APPROVED"').fetchone()[0],
        "rejected_sessions":  conn.execute('SELECT COUNT(*) FROM sessions WHERE status="REJECTED"').fetchone()[0],
        "emergency_sessions": conn.execute('SELECT COUNT(*) FROM sessions WHERE priority="emergency"').fetchone()[0],
        "total_education":    conn.execute('SELECT COUNT(*) FROM education').fetchone()[0],
        "total_events":       conn.execute('SELECT COUNT(*) FROM events').fetchone()[0],
        "total_innovations":  conn.execute('SELECT COUNT(*) FROM innovations').fetchone()[0],
        "total_achievements": conn.execute('SELECT COUNT(*) FROM achievements').fetchone()[0],
        "total_forum_posts":  conn.execute('SELECT COUNT(*) FROM forum_posts').fetchone()[0],
        "total_officers":     conn.execute('SELECT COUNT(*) FROM about_officers').fetchone()[0],
        "total_stakeholders": conn.execute('SELECT COUNT(*) FROM about_stakeholders').fetchone()[0],
    }
    conn.close()
    return api_success(stats)


# -----------------------------------------------------------------------------
# 👤 USERS — CRUD
# -----------------------------------------------------------------------------
@app.route('/api/v1/users', methods=['GET'])
@api_key_required
def api_users_list():
    """GET /api/v1/users — Daftar semua user
    Query params: role, search, limit (default 100)
    """
    role   = request.args.get('role')
    search = request.args.get('search', '')
    limit  = min(int(request.args.get('limit', 100)), 500)

    conn = get_db_connection()
    query  = 'SELECT id, username, birth_date, role, points, bio, profile_pic FROM users WHERE 1=1'
    params = []
    if role:
        query += ' AND role = ?'
        params.append(role)
    if search:
        query += ' AND username LIKE ?'
        params.append(f'%{search}%')
    query += f' ORDER BY id DESC LIMIT {limit}'
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    conn.close()
    return api_success(rows)


@app.route('/api/v1/users', methods=['POST'])
@api_key_required
def api_users_create():
    """POST /api/v1/users — Tambah user baru
    Body JSON: {username, password, birth_date, role}
    """
    data = request.get_json(silent=True) or request.form
    username   = (data.get('username') or '').strip()
    password   = (data.get('password') or '').strip()
    birth_date = (data.get('birth_date') or '2005-01-01').strip()
    role       = (data.get('role') or 'anggota_remaja').strip()

    if not username or not password:
        return api_error("username dan password wajib diisi", 400)

    conn = get_db_connection()
    existing = conn.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
    if existing:
        conn.close()
        return api_error(f"Username '{username}' sudah digunakan", 409)

    conn.execute('INSERT INTO users (username, password, birth_date, role) VALUES (?, ?, ?, ?)',
                 (username, password, birth_date, role))
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return api_success({"id": new_id, "username": username, "role": role}, "User berhasil ditambahkan", 201)


@app.route('/api/v1/users/<int:user_id>', methods=['GET'])
@api_key_required
def api_users_get(user_id):
    """GET /api/v1/users/<id> — Detail user"""
    conn = get_db_connection()
    row = conn.execute('SELECT id, username, birth_date, role, points, bio, profile_pic FROM users WHERE id = ?', (user_id,)).fetchone()
    conn.close()
    if not row:
        return api_error("User tidak ditemukan", 404)
    return api_success(dict(row))


@app.route('/api/v1/users/<int:user_id>', methods=['PUT'])
@api_key_required
def api_users_update(user_id):
    """PUT /api/v1/users/<id> — Update user
    Body JSON: {username?, password?, role?, bio?, birth_date?, points?}
    """
    data = request.get_json(silent=True) or request.form
    allowed = ['username', 'password', 'role', 'bio', 'birth_date', 'points']
    updates = {k: data[k] for k in allowed if k in data and data[k] is not None}
    if not updates:
        return api_error("Tidak ada field yang diperbarui", 400)

    conn = get_db_connection()
    existing = conn.execute('SELECT id FROM users WHERE id = ?', (user_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("User tidak ditemukan", 404)

    set_clause = ', '.join([f"{k} = ?" for k in updates.keys()])
    values = list(updates.values()) + [user_id]
    conn.execute(f'UPDATE users SET {set_clause} WHERE id = ?', values)
    conn.commit()
    conn.close()
    return api_success({"id": user_id}, "User berhasil diperbarui")


@app.route('/api/v1/users/<int:user_id>', methods=['DELETE'])
@api_key_required
def api_users_delete(user_id):
    """DELETE /api/v1/users/<id> — Hapus user"""
    conn = get_db_connection()
    existing = conn.execute('SELECT id, username FROM users WHERE id = ?', (user_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("User tidak ditemukan", 404)
    conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    return api_success({"id": user_id}, f"User '{existing['username']}' berhasil dihapus")


# -----------------------------------------------------------------------------
# 📅 SESSIONS — CRUD (Konseling)
# -----------------------------------------------------------------------------
@app.route('/api/v1/sessions', methods=['GET'])
@api_key_required
def api_sessions_list():
    """GET /api/v1/sessions — Daftar sesi konseling
    Query params: status, priority, counselor_name, member_name, limit
    """
    status   = request.args.get('status')
    priority = request.args.get('priority')
    counselor= request.args.get('counselor_name')
    member   = request.args.get('member_name')
    limit    = min(int(request.args.get('limit', 100)), 500)

    conn  = get_db_connection()
    query = 'SELECT * FROM sessions WHERE 1=1'
    params= []
    if status:
        query += ' AND UPPER(status) = UPPER(?)'
        params.append(status)
    if priority:
        query += ' AND priority = ?'
        params.append(priority)
    if counselor:
        query += ' AND counselor_name = ?'
        params.append(counselor)
    if member:
        query += ' AND member_name = ?'
        params.append(member)
    query += f' ORDER BY id DESC LIMIT {limit}'
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    conn.close()
    return api_success(rows)


@app.route('/api/v1/sessions', methods=['POST'])
@api_key_required
def api_sessions_create():
    """POST /api/v1/sessions — Tambah sesi baru
    Body JSON: {member_name, counselor_name, topic, date, time, status?}
    """
    data = request.get_json(silent=True) or request.form
    member_name   = (data.get('member_name') or '').strip()
    counselor_name= (data.get('counselor_name') or '').strip()
    topic         = (data.get('topic') or '').strip()
    date          = (data.get('date') or '').strip()
    time_val      = (data.get('time') or '').strip()
    status        = (data.get('status') or 'PENDING').strip().upper()

    if not all([member_name, counselor_name, date, time_val]):
        return api_error("member_name, counselor_name, date, dan time wajib diisi", 400)

    conn = get_db_connection()
    conflict = conn.execute('''
        SELECT 1 FROM sessions 
        WHERE counselor_name = ? AND date = ? AND time = ? AND status IN ('PENDING','APPROVED')
    ''', (counselor_name, date, time_val)).fetchone()
    if conflict:
        conn.close()
        return api_error(f"Jadwal {date} {time_val} untuk {counselor_name} sudah dibooking", 409)

    conn.execute('INSERT INTO sessions (member_name, counselor_name, topic, date, time, status) VALUES (?,?,?,?,?,?)',
                 (member_name, counselor_name, topic, date, time_val, status))
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return api_success({"id": new_id}, "Sesi konseling berhasil ditambahkan", 201)


@app.route('/api/v1/sessions/<int:session_id>', methods=['GET'])
@api_key_required
def api_sessions_get(session_id):
    """GET /api/v1/sessions/<id> — Detail sesi"""
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM sessions WHERE id = ?', (session_id,)).fetchone()
    conn.close()
    if not row:
        return api_error("Sesi tidak ditemukan", 404)
    return api_success(dict(row))


@app.route('/api/v1/sessions/<int:session_id>', methods=['PUT'])
@api_key_required
def api_sessions_update(session_id):
    """PUT /api/v1/sessions/<id> — Update sesi
    Body JSON: {status?, priority?, date?, time?, topic?}
    """
    data = request.get_json(silent=True) or request.form
    allowed = ['status', 'priority', 'date', 'time', 'topic', 'counselor_name', 'member_name']
    updates = {k: data[k] for k in allowed if k in data and data[k] is not None}
    if not updates:
        return api_error("Tidak ada field yang diperbarui", 400)

    # Normalkan status ke uppercase
    if 'status' in updates:
        updates['status'] = updates['status'].upper()

    conn = get_db_connection()
    existing = conn.execute('SELECT id FROM sessions WHERE id = ?', (session_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Sesi tidak ditemukan", 404)

    set_clause = ', '.join([f"{k} = ?" for k in updates.keys()])
    values = list(updates.values()) + [session_id]
    conn.execute(f'UPDATE sessions SET {set_clause} WHERE id = ?', values)
    conn.commit()
    conn.close()
    return api_success({"id": session_id}, "Sesi berhasil diperbarui")


@app.route('/api/v1/sessions/<int:session_id>', methods=['DELETE'])
@api_key_required
def api_sessions_delete(session_id):
    """DELETE /api/v1/sessions/<id> — Hapus sesi"""
    conn = get_db_connection()
    existing = conn.execute('SELECT id FROM sessions WHERE id = ?', (session_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Sesi tidak ditemukan", 404)
    conn.execute('DELETE FROM sessions WHERE id = ?', (session_id,))
    conn.commit()
    conn.close()
    return api_success({"id": session_id}, "Sesi berhasil dihapus")


# -----------------------------------------------------------------------------
# 📚 EDUCATION — CRUD Artikel Edukasi
# -----------------------------------------------------------------------------
@app.route('/api/v1/education', methods=['GET'])
@api_key_required
def api_education_list():
    """GET /api/v1/education — Daftar artikel edukasi"""
    limit = min(int(request.args.get('limit', 100)), 500)
    conn  = get_db_connection()
    rows  = [dict(r) for r in conn.execute('SELECT * FROM education ORDER BY id DESC LIMIT ?', (limit,)).fetchall()]
    conn.close()
    return api_success(rows)


@app.route('/api/v1/education', methods=['POST'])
@api_key_required
def api_education_create():
    """POST /api/v1/education — Tambah artikel edukasi
    Body: form-data {title, content, author, author_id?, dokumen(file)?}
    """
    title     = (request.form.get('title') or '').strip()
    content   = (request.form.get('content') or '').strip()
    author    = (request.form.get('author') or 'admin').strip()
    author_id = request.form.get('author_id', 0)

    nama_dokumen = None
    file = request.files.get('dokumen')
    if file and file.filename:
        nama_dokumen = f"{int(datetime.now().timestamp())}_{secure_filename(file.filename)}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen))

    if not title:
        return api_error("title wajib diisi", 400)
    if not content and not nama_dokumen:
        return api_error("content atau dokumen wajib ada", 400)

    conn = get_db_connection()
    conn.execute('INSERT INTO education (title, content, author, author_id, dokumen) VALUES (?,?,?,?,?)',
                 (title, content, author, author_id, nama_dokumen))
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return api_success({"id": new_id, "title": title}, "Artikel edukasi berhasil ditambahkan", 201)


@app.route('/api/v1/education/<int:edu_id>', methods=['GET'])
@api_key_required
def api_education_get(edu_id):
    """GET /api/v1/education/<id> — Detail artikel"""
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM education WHERE id = ?', (edu_id,)).fetchone()
    if not row:
        conn.close()
        return api_error("Artikel tidak ditemukan", 404)
    article = dict(row)
    article['comments'] = [dict(c) for c in conn.execute(
        'SELECT * FROM education_comments WHERE article_id = ?', (edu_id,)).fetchall()]
    conn.close()
    return api_success(article)


@app.route('/api/v1/education/<int:edu_id>', methods=['PUT'])
@api_key_required
def api_education_update(edu_id):
    """PUT /api/v1/education/<id> — Edit artikel
    Body: form-data {title?, content?, dokumen(file)?}
    """
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM education WHERE id = ?', (edu_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Artikel tidak ditemukan", 404)

    data    = request.get_json(silent=True) or {}
    title   = (request.form.get('title') or data.get('title') or existing['title']).strip()
    content = request.form.get('content') or data.get('content') or existing['content']

    nama_dokumen = existing['dokumen']
    file = request.files.get('dokumen')
    if file and file.filename:
        if nama_dokumen:
            old_path = os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen)
            if os.path.exists(old_path):
                try: os.remove(old_path)
                except: pass
        nama_dokumen = f"{int(datetime.now().timestamp())}_{secure_filename(file.filename)}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen))

    conn.execute('UPDATE education SET title = ?, content = ?, dokumen = ? WHERE id = ?',
                 (title, content, nama_dokumen, edu_id))
    conn.commit()
    conn.close()
    return api_success({"id": edu_id}, "Artikel berhasil diperbarui")


@app.route('/api/v1/education/<int:edu_id>', methods=['DELETE'])
@api_key_required
def api_education_delete(edu_id):
    """DELETE /api/v1/education/<id> — Hapus artikel"""
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM education WHERE id = ?', (edu_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Artikel tidak ditemukan", 404)
    if existing['dokumen']:
        path = os.path.join(app.config['UPLOAD_FOLDER'], existing['dokumen'])
        if os.path.exists(path):
            try: os.remove(path)
            except: pass
    conn.execute('DELETE FROM education WHERE id = ?', (edu_id,))
    conn.commit()
    conn.close()
    return api_success({"id": edu_id}, f"Artikel '{existing['title']}' berhasil dihapus")


# -----------------------------------------------------------------------------
# 🎉 EVENTS — CRUD
# -----------------------------------------------------------------------------
@app.route('/api/v1/events', methods=['GET'])
@api_key_required
def api_events_list():
    """GET /api/v1/events — Daftar events"""
    limit = min(int(request.args.get('limit', 100)), 500)
    conn  = get_db_connection()
    events_raw = conn.execute('SELECT * FROM events ORDER BY date ASC LIMIT ?', (limit,)).fetchall()
    result = []
    for e in events_raw:
        ev = dict(e)
        ev['participant_count'] = conn.execute(
            'SELECT COUNT(*) FROM event_participants WHERE event_id = ?', (e['id'],)).fetchone()[0]
        result.append(ev)
    conn.close()
    return api_success(result)


@app.route('/api/v1/events', methods=['POST'])
@api_key_required
def api_events_create():
    """POST /api/v1/events — Tambah event baru
    Body JSON: {title, description, date, time, author?}
    """
    data  = request.get_json(silent=True) or request.form
    title = (data.get('title') or '').strip()
    desc  = (data.get('description') or '').strip()
    date  = (data.get('date') or '').strip()
    t     = (data.get('time') or '').strip()
    author= (data.get('author') or 'admin').strip()

    if not title or not date:
        return api_error("title dan date wajib diisi", 400)

    conn = get_db_connection()
    conn.execute('INSERT INTO events (title, description, date, time, author) VALUES (?,?,?,?,?)',
                 (title, desc, date, t, author))
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return api_success({"id": new_id, "title": title}, "Event berhasil ditambahkan", 201)


@app.route('/api/v1/events/<int:event_id>', methods=['GET'])
@api_key_required
def api_events_get(event_id):
    """GET /api/v1/events/<id> — Detail event"""
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    if not row:
        conn.close()
        return api_error("Event tidak ditemukan", 404)
    ev = dict(row)
    ev['participants'] = [r['username'] for r in conn.execute(
        'SELECT username FROM event_participants WHERE event_id = ?', (event_id,)).fetchall()]
    conn.close()
    return api_success(ev)


@app.route('/api/v1/events/<int:event_id>', methods=['PUT'])
@api_key_required
def api_events_update(event_id):
    """PUT /api/v1/events/<id> — Update event
    Body JSON: {title?, description?, date?, time?}
    """
    data = request.get_json(silent=True) or request.form
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Event tidak ditemukan", 404)

    title = (data.get('title') or existing['title']).strip()
    desc  = data.get('description') if data.get('description') is not None else existing['description']
    date  = (data.get('date') or existing['date']).strip()
    t     = data.get('time') if data.get('time') is not None else existing['time']

    conn.execute('UPDATE events SET title=?, description=?, date=?, time=? WHERE id=?',
                 (title, desc, date, t, event_id))
    conn.commit()
    conn.close()
    return api_success({"id": event_id}, "Event berhasil diperbarui")


@app.route('/api/v1/events/<int:event_id>', methods=['DELETE'])
@api_key_required
def api_events_delete(event_id):
    """DELETE /api/v1/events/<id> — Hapus event"""
    conn = get_db_connection()
    existing = conn.execute('SELECT id, title FROM events WHERE id = ?', (event_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Event tidak ditemukan", 404)
    conn.execute('DELETE FROM event_participants WHERE event_id = ?', (event_id,))
    conn.execute('DELETE FROM events WHERE id = ?', (event_id,))
    conn.commit()
    conn.close()
    return api_success({"id": event_id}, f"Event '{existing['title']}' berhasil dihapus")


# -----------------------------------------------------------------------------
# 💡 INNOVATIONS — CRUD Inovasi & Program Kerja
# -----------------------------------------------------------------------------
@app.route('/api/v1/innovations', methods=['GET'])
@api_key_required
def api_innovations_list():
    """GET /api/v1/innovations — Daftar inovasi"""
    limit = min(int(request.args.get('limit', 100)), 500)
    conn  = get_db_connection()
    rows  = [dict(r) for r in conn.execute('SELECT * FROM innovations ORDER BY id DESC LIMIT ?', (limit,)).fetchall()]
    conn.close()
    return api_success(rows)


@app.route('/api/v1/innovations', methods=['POST'])
@api_key_required
def api_innovations_create():
    """POST /api/v1/innovations — Tambah inovasi
    Body: form-data {title, content?, author, author_id?, dokumen(file)?}
    """
    title     = (request.form.get('title') or '').strip()
    content   = (request.form.get('content') or '').strip()
    author    = (request.form.get('author') or 'admin').strip()
    author_id = request.form.get('author_id', 0)

    if not title:
        return api_error("title wajib diisi", 400)

    nama_dokumen = None
    file = request.files.get('dokumen')
    if file and file.filename:
        nama_dokumen = f"{int(datetime.now().timestamp())}_{secure_filename(file.filename)}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen))

    if not content and not nama_dokumen:
        return api_error("content atau dokumen wajib ada", 400)

    conn = get_db_connection()
    conn.execute('INSERT INTO innovations (title, content, author, author_id, dokumen) VALUES (?,?,?,?,?)',
                 (title, content, author, author_id, nama_dokumen))
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return api_success({"id": new_id, "title": title}, "Inovasi berhasil ditambahkan", 201)


@app.route('/api/v1/innovations/<int:inn_id>', methods=['GET'])
@api_key_required
def api_innovations_get(inn_id):
    """GET /api/v1/innovations/<id> — Detail inovasi"""
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM innovations WHERE id = ?', (inn_id,)).fetchone()
    if not row:
        conn.close()
        return api_error("Inovasi tidak ditemukan", 404)
    inn = dict(row)
    inn['comments'] = [dict(c) for c in conn.execute(
        'SELECT * FROM innovations_comments WHERE innovation_id = ?', (inn_id,)).fetchall()]
    conn.close()
    return api_success(inn)


@app.route('/api/v1/innovations/<int:inn_id>', methods=['PUT'])
@api_key_required
def api_innovations_update(inn_id):
    """PUT /api/v1/innovations/<id> — Update inovasi"""
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM innovations WHERE id = ?', (inn_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Inovasi tidak ditemukan", 404)

    data    = request.get_json(silent=True) or {}
    title   = (request.form.get('title') or data.get('title') or existing['title']).strip()
    content = request.form.get('content') or data.get('content') or existing['content']

    nama_dokumen = existing['dokumen']
    file = request.files.get('dokumen')
    if file and file.filename:
        if nama_dokumen:
            old_path = os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen)
            if os.path.exists(old_path):
                try: os.remove(old_path)
                except: pass
        nama_dokumen = f"{int(datetime.now().timestamp())}_{secure_filename(file.filename)}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen))

    conn.execute('UPDATE innovations SET title=?, content=?, dokumen=? WHERE id=?',
                 (title, content, nama_dokumen, inn_id))
    conn.commit()
    conn.close()
    return api_success({"id": inn_id}, "Inovasi berhasil diperbarui")


@app.route('/api/v1/innovations/<int:inn_id>', methods=['DELETE'])
@api_key_required
def api_innovations_delete(inn_id):
    """DELETE /api/v1/innovations/<id> — Hapus inovasi"""
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM innovations WHERE id = ?', (inn_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Inovasi tidak ditemukan", 404)
    if existing['dokumen']:
        path = os.path.join(app.config['UPLOAD_FOLDER'], existing['dokumen'])
        if os.path.exists(path):
            try: os.remove(path)
            except: pass
    conn.execute('DELETE FROM innovations_comments WHERE innovation_id = ?', (inn_id,))
    conn.execute('DELETE FROM innovations WHERE id = ?', (inn_id,))
    conn.commit()
    conn.close()
    return api_success({"id": inn_id}, f"Inovasi '{existing['title']}' berhasil dihapus")


# -----------------------------------------------------------------------------
# 🏆 ACHIEVEMENTS — CRUD Prestasi
# -----------------------------------------------------------------------------
@app.route('/api/v1/achievements', methods=['GET'])
@api_key_required
def api_achievements_list():
    """GET /api/v1/achievements — Daftar prestasi"""
    limit = min(int(request.args.get('limit', 100)), 500)
    conn  = get_db_connection()
    rows  = [dict(r) for r in conn.execute('SELECT * FROM achievements ORDER BY id DESC LIMIT ?', (limit,)).fetchall()]
    conn.close()
    return api_success(rows)


@app.route('/api/v1/achievements', methods=['POST'])
@api_key_required
def api_achievements_create():
    """POST /api/v1/achievements — Tambah prestasi
    Body: form-data {title, content?, author, author_id?, dokumen(file)?}
    """
    title     = (request.form.get('title') or '').strip()
    content   = (request.form.get('content') or '').strip()
    author    = (request.form.get('author') or 'admin').strip()
    author_id = request.form.get('author_id', 0)

    if not title:
        return api_error("title wajib diisi", 400)

    nama_dokumen = None
    file = request.files.get('dokumen')
    if file and file.filename:
        nama_dokumen = f"{int(datetime.now().timestamp())}_{secure_filename(file.filename)}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen).replace('\\', '/')
        file.save(filepath)

    if not content and not nama_dokumen:
        return api_error("content atau dokumen wajib ada", 400)

    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    conn = get_db_connection()
    conn.execute('INSERT INTO achievements (title, content, author, author_id, dokumen, date) VALUES (?,?,?,?,?,?)',
                 (title, content, author, author_id, nama_dokumen, date_str))
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return api_success({"id": new_id, "title": title}, "Prestasi berhasil ditambahkan", 201)


@app.route('/api/v1/achievements/<int:ach_id>', methods=['GET'])
@api_key_required
def api_achievements_get(ach_id):
    """GET /api/v1/achievements/<id> — Detail prestasi"""
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM achievements WHERE id = ?', (ach_id,)).fetchone()
    if not row:
        conn.close()
        return api_error("Prestasi tidak ditemukan", 404)
    ach = dict(row)
    ach['comments'] = [dict(c) for c in conn.execute(
        'SELECT * FROM achievements_comments WHERE achievement_id = ?', (ach_id,)).fetchall()]
    conn.close()
    return api_success(ach)


@app.route('/api/v1/achievements/<int:ach_id>', methods=['PUT'])
@api_key_required
def api_achievements_update(ach_id):
    """PUT /api/v1/achievements/<id> — Update prestasi"""
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM achievements WHERE id = ?', (ach_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Prestasi tidak ditemukan", 404)

    data    = request.get_json(silent=True) or {}
    title   = (request.form.get('title') or data.get('title') or existing['title']).strip()
    content = request.form.get('content') or data.get('content') or existing['content']

    nama_dokumen = existing['dokumen']
    file = request.files.get('dokumen')
    if file and file.filename:
        if nama_dokumen:
            old_path = os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen)
            if os.path.exists(old_path):
                try: os.remove(old_path)
                except: pass
        nama_dokumen = f"{int(datetime.now().timestamp())}_{secure_filename(file.filename)}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_dokumen).replace('\\', '/'))

    conn.execute('UPDATE achievements SET title=?, content=?, dokumen=? WHERE id=?',
                 (title, content, nama_dokumen, ach_id))
    conn.commit()
    conn.close()
    return api_success({"id": ach_id}, "Prestasi berhasil diperbarui")


@app.route('/api/v1/achievements/<int:ach_id>', methods=['DELETE'])
@api_key_required
def api_achievements_delete(ach_id):
    """DELETE /api/v1/achievements/<id> — Hapus prestasi"""
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM achievements WHERE id = ?', (ach_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Prestasi tidak ditemukan", 404)
    if existing['dokumen']:
        path = os.path.join(app.config['UPLOAD_FOLDER'], existing['dokumen'])
        if os.path.exists(path):
            try: os.remove(path)
            except: pass
    conn.execute('DELETE FROM achievements_comments WHERE achievement_id = ?', (ach_id,))
    conn.execute('DELETE FROM achievements WHERE id = ?', (ach_id,))
    conn.commit()
    conn.close()
    return api_success({"id": ach_id}, f"Prestasi '{existing['title']}' berhasil dihapus")


# -----------------------------------------------------------------------------
# 📢 FORUM — CRUD Postingan Forum
# -----------------------------------------------------------------------------
@app.route('/api/v1/forum', methods=['GET'])
@api_key_required
def api_forum_list():
    """GET /api/v1/forum — Daftar postingan forum
    Query params: is_announcement (1/0), limit
    """
    is_ann = request.args.get('is_announcement')
    limit  = min(int(request.args.get('limit', 100)), 500)
    conn   = get_db_connection()
    query  = 'SELECT * FROM forum_posts WHERE 1=1'
    params = []
    if is_ann is not None:
        query += ' AND is_announcement = ?'
        params.append(int(is_ann))
    query += f' ORDER BY is_announcement DESC, id DESC LIMIT {limit}'
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    conn.close()
    return api_success(rows)


@app.route('/api/v1/forum', methods=['POST'])
@api_key_required
def api_forum_create():
    """POST /api/v1/forum — Buat postingan forum
    Body JSON: {username, content, role?, is_announcement?}
    """
    data    = request.get_json(silent=True) or request.form
    username= (data.get('username') or 'admin').strip()
    content = (data.get('content') or '').strip()
    role    = (data.get('role') or 'admin_pikr').strip()
    is_ann  = int(data.get('is_announcement', 0))

    if not content:
        return api_error("content wajib diisi", 400)

    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    conn = get_db_connection()
    conn.execute('INSERT INTO forum_posts (username, content, role, is_announcement, date) VALUES (?,?,?,?,?)',
                 (username, content, role, is_ann, date_str))
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return api_success({"id": new_id}, "Postingan forum berhasil dibuat", 201)


@app.route('/api/v1/forum/<int:post_id>', methods=['GET'])
@api_key_required
def api_forum_get(post_id):
    """GET /api/v1/forum/<id> — Detail postingan"""
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM forum_posts WHERE id = ?', (post_id,)).fetchone()
    conn.close()
    if not row:
        return api_error("Postingan tidak ditemukan", 404)
    return api_success(dict(row))


@app.route('/api/v1/forum/<int:post_id>', methods=['PUT'])
@api_key_required
def api_forum_update(post_id):
    """PUT /api/v1/forum/<id> — Edit postingan
    Body JSON: {content?, is_announcement?}
    """
    data = request.get_json(silent=True) or request.form
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM forum_posts WHERE id = ?', (post_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Postingan tidak ditemukan", 404)

    content = data.get('content') if data.get('content') is not None else existing['content']
    is_ann  = int(data.get('is_announcement')) if data.get('is_announcement') is not None else existing['is_announcement']

    conn.execute('UPDATE forum_posts SET content=?, is_announcement=? WHERE id=?',
                 (content, is_ann, post_id))
    conn.commit()
    conn.close()
    return api_success({"id": post_id}, "Postingan berhasil diperbarui")


@app.route('/api/v1/forum/<int:post_id>', methods=['DELETE'])
@api_key_required
def api_forum_delete(post_id):
    """DELETE /api/v1/forum/<id> — Hapus postingan"""
    conn = get_db_connection()
    existing = conn.execute('SELECT id FROM forum_posts WHERE id = ?', (post_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Postingan tidak ditemukan", 404)
    conn.execute('DELETE FROM forum_posts WHERE id = ?', (post_id,))
    conn.commit()
    conn.close()
    return api_success({"id": post_id}, "Postingan berhasil dihapus")


# -----------------------------------------------------------------------------
# 🏛️ TENTANG KAMI — Pengurus
# -----------------------------------------------------------------------------
@app.route('/api/v1/about/officers', methods=['GET'])
@api_key_required
def api_officers_list():
    """GET /api/v1/about/officers — Daftar pengurus"""
    conn = get_db_connection()
    rows = [dict(r) for r in conn.execute('SELECT * FROM about_officers ORDER BY id ASC').fetchall()]
    conn.close()
    return api_success(rows)


@app.route('/api/v1/about/officers', methods=['POST'])
@api_key_required
def api_officers_create():
    """POST /api/v1/about/officers — Tambah pengurus
    Body: form-data {name, position, caption?, photo(file)}
    """
    name     = (request.form.get('name') or '').strip()
    position = (request.form.get('position') or '').strip()
    caption  = (request.form.get('caption') or '').strip()

    if not name or not position:
        return api_error("name dan position wajib diisi", 400)

    photo_path = ''
    file = request.files.get('photo')
    if file and file.filename:
        filename = f"officer_{int(time.time())}_{secure_filename(file.filename)}"
        filepath = os.path.join('static/uploads/about', filename).replace('\\', '/')
        os.makedirs('static/uploads/about', exist_ok=True)
        file.save(filepath)
        photo_path = filepath
    else:
        return api_error("photo (file) wajib diunggah", 400)

    conn = get_db_connection()
    conn.execute('INSERT INTO about_officers (name, position, photo, caption) VALUES (?,?,?,?)',
                 (name, position, photo_path, caption))
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return api_success({"id": new_id, "name": name}, "Pengurus berhasil ditambahkan", 201)


@app.route('/api/v1/about/officers/<int:officer_id>', methods=['GET'])
@api_key_required
def api_officers_get(officer_id):
    """GET /api/v1/about/officers/<id> — Detail pengurus"""
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM about_officers WHERE id = ?', (officer_id,)).fetchone()
    conn.close()
    if not row:
        return api_error("Pengurus tidak ditemukan", 404)
    return api_success(dict(row))


@app.route('/api/v1/about/officers/<int:officer_id>', methods=['PUT'])
@api_key_required
def api_officers_update(officer_id):
    """PUT /api/v1/about/officers/<id> — Update pengurus
    Body: form-data {name?, position?, caption?, photo(file)?}
    """
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM about_officers WHERE id = ?', (officer_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Pengurus tidak ditemukan", 404)

    name     = (request.form.get('name') or existing['name']).strip()
    position = (request.form.get('position') or existing['position']).strip()
    caption  = request.form.get('caption') if request.form.get('caption') is not None else (existing['caption'] or '')

    photo_path = existing['photo']
    file = request.files.get('photo')
    if file and file.filename:
        if existing['photo'] and os.path.exists(existing['photo']):
            try: os.remove(existing['photo'])
            except: pass
        filename = f"officer_{int(time.time())}_{secure_filename(file.filename)}"
        filepath = os.path.join('static/uploads/about', filename).replace('\\', '/')
        os.makedirs('static/uploads/about', exist_ok=True)
        file.save(filepath)
        photo_path = filepath

    conn.execute('UPDATE about_officers SET name=?, position=?, photo=?, caption=? WHERE id=?',
                 (name, position, photo_path, caption, officer_id))
    conn.commit()
    conn.close()
    return api_success({"id": officer_id}, "Data pengurus berhasil diperbarui")


@app.route('/api/v1/about/officers/<int:officer_id>', methods=['DELETE'])
@api_key_required
def api_officers_delete(officer_id):
    """DELETE /api/v1/about/officers/<id> — Hapus pengurus"""
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM about_officers WHERE id = ?', (officer_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Pengurus tidak ditemukan", 404)
    if existing['photo'] and os.path.exists(existing['photo']):
        try: os.remove(existing['photo'])
        except: pass
    conn.execute('DELETE FROM about_officers WHERE id = ?', (officer_id,))
    conn.commit()
    conn.close()
    return api_success({"id": officer_id}, f"Pengurus '{existing['name']}' berhasil dihapus")


# -----------------------------------------------------------------------------
# 🏛️ TENTANG KAMI — Stakeholder
# -----------------------------------------------------------------------------
@app.route('/api/v1/about/stakeholders', methods=['GET'])
@api_key_required
def api_stakeholders_list():
    """GET /api/v1/about/stakeholders — Daftar stakeholder"""
    conn = get_db_connection()
    rows = [dict(r) for r in conn.execute('SELECT * FROM about_stakeholders ORDER BY id ASC').fetchall()]
    conn.close()
    return api_success(rows)


@app.route('/api/v1/about/stakeholders', methods=['POST'])
@api_key_required
def api_stakeholders_create():
    """POST /api/v1/about/stakeholders — Tambah stakeholder
    Body: form-data {name, role_title, role_desc?, category?, icon_type?, logo(file)?}
    """
    name       = (request.form.get('name') or '').strip()
    role_title = (request.form.get('role_title') or '').strip()
    role_desc  = (request.form.get('role_desc') or '').strip()
    category   = (request.form.get('category') or '').strip()
    icon_type  = (request.form.get('icon_type') or 'community').strip()

    if not name or not role_title:
        return api_error("name dan role_title wajib diisi", 400)

    logo_filename = None
    file = request.files.get('logo')
    if file and file.filename:
        os.makedirs('static/uploads/stakeholders', exist_ok=True)
        logo_filename = f"logo_{int(time.time())}_{secure_filename(file.filename)}"
        file.save(os.path.join('static/uploads/stakeholders', logo_filename))

    conn = get_db_connection()
    conn.execute('INSERT INTO about_stakeholders (name, role_title, role_desc, category, icon_type, logo) VALUES (?,?,?,?,?,?)',
                 (name, role_title, role_desc, category, icon_type, logo_filename))
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return api_success({"id": new_id, "name": name}, "Stakeholder berhasil ditambahkan", 201)


@app.route('/api/v1/about/stakeholders/<int:stk_id>', methods=['GET'])
@api_key_required
def api_stakeholders_get(stk_id):
    """GET /api/v1/about/stakeholders/<id> — Detail stakeholder"""
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM about_stakeholders WHERE id = ?', (stk_id,)).fetchone()
    conn.close()
    if not row:
        return api_error("Stakeholder tidak ditemukan", 404)
    return api_success(dict(row))


@app.route('/api/v1/about/stakeholders/<int:stk_id>', methods=['PUT'])
@api_key_required
def api_stakeholders_update(stk_id):
    """PUT /api/v1/about/stakeholders/<id> — Update stakeholder"""
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM about_stakeholders WHERE id = ?', (stk_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Stakeholder tidak ditemukan", 404)

    name       = (request.form.get('name') or existing['name']).strip()
    role_title = (request.form.get('role_title') or existing['role_title']).strip()
    role_desc  = request.form.get('role_desc') if request.form.get('role_desc') is not None else (existing['role_desc'] or '')
    category   = request.form.get('category') if request.form.get('category') is not None else (existing['category'] or '')
    icon_type  = (request.form.get('icon_type') or existing['icon_type'] or 'community').strip()

    logo_filename = existing['logo']
    file = request.files.get('logo')
    if file and file.filename:
        if logo_filename:
            old_path = os.path.join('static/uploads/stakeholders', logo_filename)
            if os.path.exists(old_path):
                try: os.remove(old_path)
                except: pass
        os.makedirs('static/uploads/stakeholders', exist_ok=True)
        logo_filename = f"logo_{int(time.time())}_{secure_filename(file.filename)}"
        file.save(os.path.join('static/uploads/stakeholders', logo_filename))

    conn.execute('UPDATE about_stakeholders SET name=?, role_title=?, role_desc=?, category=?, icon_type=?, logo=? WHERE id=?',
                 (name, role_title, role_desc, category, icon_type, logo_filename, stk_id))
    conn.commit()
    conn.close()
    return api_success({"id": stk_id}, "Stakeholder berhasil diperbarui")


@app.route('/api/v1/about/stakeholders/<int:stk_id>', methods=['DELETE'])
@api_key_required
def api_stakeholders_delete(stk_id):
    """DELETE /api/v1/about/stakeholders/<id> — Hapus stakeholder"""
    conn = get_db_connection()
    existing = conn.execute('SELECT * FROM about_stakeholders WHERE id = ?', (stk_id,)).fetchone()
    if not existing:
        conn.close()
        return api_error("Stakeholder tidak ditemukan", 404)
    if existing['logo']:
        logo_path = os.path.join('static/uploads/stakeholders', existing['logo'])
        if os.path.exists(logo_path):
            try: os.remove(logo_path)
            except: pass
    conn.execute('DELETE FROM about_stakeholders WHERE id = ?', (stk_id,))
    conn.commit()
    conn.close()
    return api_success({"id": stk_id}, f"Stakeholder '{existing['name']}' berhasil dihapus")


# -----------------------------------------------------------------------------
# 🔍 API DOCS — Daftar semua endpoint
# -----------------------------------------------------------------------------
@app.route('/api/v1', methods=['GET'])
@app.route('/api/v1/', methods=['GET'])
def api_docs():
    """GET /api/v1/ — Dokumentasi singkat endpoint API"""
    endpoints = {
        "version": "1.0",
        "auth": "Tambahkan header 'X-API-Key: <api_key>' di setiap request",
        "endpoints": {
            "stats":        {"GET": "/api/v1/stats"},
            "users":        {"GET": "/api/v1/users", "POST": "/api/v1/users",
                             "GET_ONE": "/api/v1/users/<id>", "PUT": "/api/v1/users/<id>",
                             "DELETE": "/api/v1/users/<id>"},
            "sessions":     {"GET": "/api/v1/sessions", "POST": "/api/v1/sessions",
                             "GET_ONE": "/api/v1/sessions/<id>", "PUT": "/api/v1/sessions/<id>",
                             "DELETE": "/api/v1/sessions/<id>"},
            "education":    {"GET": "/api/v1/education", "POST": "/api/v1/education",
                             "GET_ONE": "/api/v1/education/<id>", "PUT": "/api/v1/education/<id>",
                             "DELETE": "/api/v1/education/<id>"},
            "events":       {"GET": "/api/v1/events", "POST": "/api/v1/events",
                             "GET_ONE": "/api/v1/events/<id>", "PUT": "/api/v1/events/<id>",
                             "DELETE": "/api/v1/events/<id>"},
            "innovations":  {"GET": "/api/v1/innovations", "POST": "/api/v1/innovations",
                             "GET_ONE": "/api/v1/innovations/<id>", "PUT": "/api/v1/innovations/<id>",
                             "DELETE": "/api/v1/innovations/<id>"},
            "achievements": {"GET": "/api/v1/achievements", "POST": "/api/v1/achievements",
                             "GET_ONE": "/api/v1/achievements/<id>", "PUT": "/api/v1/achievements/<id>",
                             "DELETE": "/api/v1/achievements/<id>"},
            "forum":        {"GET": "/api/v1/forum", "POST": "/api/v1/forum",
                             "GET_ONE": "/api/v1/forum/<id>", "PUT": "/api/v1/forum/<id>",
                             "DELETE": "/api/v1/forum/<id>"},
            "officers":     {"GET": "/api/v1/about/officers", "POST": "/api/v1/about/officers",
                             "GET_ONE": "/api/v1/about/officers/<id>", "PUT": "/api/v1/about/officers/<id>",
                             "DELETE": "/api/v1/about/officers/<id>"},
            "stakeholders": {"GET": "/api/v1/about/stakeholders", "POST": "/api/v1/about/stakeholders",
                             "GET_ONE": "/api/v1/about/stakeholders/<id>", "PUT": "/api/v1/about/stakeholders/<id>",
                             "DELETE": "/api/v1/about/stakeholders/<id>"},
        }
    }
    return jsonify(endpoints)


if __name__ == '__main__':
    app.run(debug=True)